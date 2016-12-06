""" Io

:Author: Jonathan Karr <karr@mssm.edu>
:Date: 2016-11-23
:Copyright: 2016, Karr Lab
:License: MIT
"""

from collections import OrderedDict
from itertools import chain
from natsort import natsorted, ns
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from wc_utils.util.list import transpose
from wc_utils.schema import utils
from wc_utils.schema.core import Model, Attribute, RelatedAttribute, Validator, TabularOrientation
from six import integer_types, string_types


class ExcelIo(object):

    @classmethod
    def write(cls, filename, objects, model_order):
        """ Write a set of model objects to an Excel workbook with one worksheet for each `Model`

        Args:
            filename (:obj:`str`): path to write Excel file
            objects (:obj:`set`): set of objects
            model_order (:obj:`list`): list of model, in the order that they should
                appear as worksheets; all models which are not in `model_order` will
                follow in alphabetical order
        """

        # get related objects
        more_objects = set()
        for obj in objects:
            more_objects.update(obj.get_related())

        # clean objects
        all_objects = objects | more_objects
        error = Validator().run(all_objects)

        if error:
            raise ValueError(str(error))

        # group objects by class
        grouped_objects = {}
        for obj in all_objects:
            if obj.__class__ not in grouped_objects:
                grouped_objects[obj.__class__] = set()
            grouped_objects[obj.__class__].add(obj)

        # create workbook
        workbook = Workbook()

        # remove default sheet
        workbook.remove_sheet(workbook.active)

        # add sheets
        unordered_models = natsorted(set(grouped_objects.keys()).difference(set(model_order)),
                                     lambda model: model.Meta.verbose_name_plural, alg=ns.IGNORECASE)

        for model in chain(model_order, unordered_models):
            if model.Meta.tabular_orientation == TabularOrientation['inline']:
                continue

            if model in grouped_objects:
                objects = grouped_objects[model]
            else:
                objects = set()

            cls.write_model(workbook, model, objects)

        # save workbook
        workbook.save(filename)

    @classmethod
    def write_model(cls, workbook, model, objects):
        """ Write a set of model objects to an Excel worksheet

        Args:
            workbook (:obj:`Workbook`): workbook
            model (:obj:`class`): model
            objects (:obj:`set` of `Model`): set of instances of `model`
        """

        # attribute order
        attributes = [model.Meta.attributes[attr_name] for attr_name in model.Meta.attribute_order]

        # column labels
        headings = [[attr.verbose_name for attr in attributes]]

        # objects
        objects = natsorted(objects, lambda obj: obj.serialize(), alg=ns.IGNORECASE)

        data = []
        for obj in objects:
            obj_data = []
            for attr in attributes:
                obj_data.append(attr.serialize(getattr(obj, attr.name)))
            data.append(obj_data)

        # transpose data for column orientation
        if model.Meta.tabular_orientation == TabularOrientation['row']:
            cls.write_sheet(workbook,
                            sheet_name=model.Meta.verbose_name_plural,
                            data=data,
                            column_headings=headings,
                            frozen_rows=1,
                            frozen_columns=model.Meta.frozen_columns,
                            )
        else:
            cls.write_sheet(workbook,
                            sheet_name=model.Meta.verbose_name_plural,
                            data=transpose(data),
                            row_headings=headings,
                            frozen_rows=model.Meta.frozen_columns,
                            frozen_columns=1,
                            )

    @staticmethod
    def write_sheet(workbook, sheet_name, data,
                    row_headings=None, column_headings=None,
                    frozen_rows=0, frozen_columns=0):
        """ Write data to sheet

        Args:
            workbook (:obj:`Workbook`): workbook
            sheet_name (:obj:`str`): sheet name
            data (:obj:`list` of `list` of `object`): list of list of cell values
            row_headings (:obj:`list` of `list` of `str`, optional): list of list of row headings
            column_headings (:obj:`list` of `list` of `str`, optional): list of list of column headings
            frozen_rows (:obj:`int`, optional): number of rows to freeze
            frozen_columns (:obj:`int`, optional): number of columns to freeze
        """
        row_headings = row_headings or []
        column_headings = column_headings or []

        # create sheet
        ws = workbook.create_sheet(sheet_name)

        # initialize styling
        ws.freeze_panes = ws.cell(row=frozen_rows + 1, column=frozen_columns + 1)

        alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        fill = PatternFill("solid", fgColor='CCCCCC')
        font = Font(bold=True)
        height = 15

        # row headings
        for i_col, row_headings_col in enumerate(row_headings):
            for i_row, heading in enumerate(row_headings_col):
                cell = ws.cell(row=1 + len(column_headings) + i_row, column=1 + i_col)
                cell.value = heading
                cell.alignment = alignment
                cell.fill = fill
                cell.font = font

        # column headings
        for i_row, column_headings_row in enumerate(column_headings):
            for i_col, heading in enumerate(column_headings_row):
                cell = ws.cell(row=1 + i_row, column=1 + len(row_headings) + i_col)
                cell.value = heading
                cell.alignment = alignment
                cell.fill = fill
                cell.font = font

        # data
        for i_row, data_row in enumerate(data):
            for i_col, value in enumerate(data_row):
                cell = ws.cell(row=1 + len(column_headings) + i_row, column=1 + len(row_headings) + i_col)

                if isinstance(value, string_types):
                    data_type = Cell.TYPE_STRING
                elif isinstance(value, bool):
                    data_type = Cell.TYPE_BOOL
                elif isinstance(value, (integer_types, float)):
                    data_type = Cell.TYPE_NUMERIC
                elif value is not None:
                    raise ValueError('Cannot save values of type "{}" at sheet "{}" row={}, attribute={}'.format(
                        value.__class__.__name__,
                        sheet_name,
                        1 + len(column_headings) + i_row,
                        column_headings[-1][i_col]))

                if value is not None:
                    cell.set_explicit_value(value=value, data_type=data_type)
                cell.alignment = alignment

        # row heights
        for i_row in range(1, ws.max_row + 1):
            ws.row_dimensions[i_row].height = height

    @classmethod
    def read(cls, filename, models):
        """ Read a set of model objects from an Excel workbook

        Args:
            filename (:obj:`str`): path to Excel worksheet
            models (:obj:`set` of `class`): set of models

        Returns:
            :obj:`dict`: model objects grouped by `Model`
        """
        # load workbook
        workbook = load_workbook(filename=filename)

        # check that models are defined for each worksheet
        sheet_names = set(workbook.get_sheet_names())
        model_names = set((model.Meta.verbose_name_plural for model in models))
        extra_sheets = sheet_names.difference(model_names)
        if extra_sheets:
            raise ValueError('Models must be defined for the following worksheets: {}'.format(', '.join(extra_sheets)))

        # read objects
        attributes = {}
        data = {}
        errors = {}
        objects = {}
        for model in models:
            model_attributes, model_data, model_errors, model_objects = cls.read_model(workbook, model)
            if model_attributes:
                attributes[model] = model_attributes
            if model_data:
                data[model] = model_data
            if model_errors:
                errors[model] = model_errors
            if model_objects:
                objects[model] = model_objects

        if errors:
            msg = 'The model cannot be loaded because the spreadsheet contains error(s):\n'
            for model, model_errors in errors.items():
                msg += '- {}:\n  - {}\n'.format(model.__name__, '\n  - '.join(model_errors))
            raise ValueError(msg)

        # link objects
        objects_by_primary_attribute = {}
        for model, objects_model in objects.items():
            objects_by_primary_attribute[model] = {obj.get_primary_attribute(): obj for obj in objects_model}

        errors = {}
        for model, objects_model in objects.items():
            model_errors = cls.link_model(model, attributes[model], data[model], objects_model,
                                          objects_by_primary_attribute)
            if model_errors:
                errors[model] = model_errors

        if errors:
            msg = 'The model cannot be loaded because the spreadsheet contains error(s):\n'
            for model, model_errors in errors.items():
                msg += '- {}:\n'.format(model.__name__)
                for model_error in model_errors:
                    if isinstance(model_error, str):
                        msg += '  - {}\n'.format(model_error)
                    else:
                        msg += '  - {}\n'.format('  - {}\n'.join(model_error.messages))
            raise ValueError(msg)

        # convert to sets
        for model in models:
            if model in objects:
                objects[model] = set(objects[model])
            else:
                objects[model] = set()

        for model, model_objects in objects_by_primary_attribute.items():
            if model not in objects:
                objects[model] = set()
            objects[model].update(model_objects.values())

        # validate
        all_objects = set()
        for model in models:
            all_objects.update(objects[model])

        errors = Validator().run(all_objects)
        if errors:
            raise ValueError(str(errors))

        # return
        return objects

    @classmethod
    def read_model(cls, workbook, model):
        """ Read a set of objects from an Excel worksheet

        Args:
            workbook (:obj:`Workbook`): workbook
            model (:obj:`class`): model

        Returns:
            :obj:`tuple` of
                `list` of `Attribute`,
                `list` of `list` of `object`,
                `list` of `str`,
                `list` of `Model`: tuple of
                * attribute order of `data`
                * a two-dimensional nested list of object data
                * a list of parsing errors
                * constructed model objects
        """
        if model.Meta.verbose_name_plural not in workbook:
            return ([], [], None, [])

        # get worksheet
        if model.Meta.tabular_orientation == TabularOrientation['row']:
            data, _, headings = cls.read_sheet(
                workbook,
                model.Meta.verbose_name_plural,
                num_column_heading_rows=1)
        else:
            data, headings, _ = cls.read_sheet(
                workbook,
                model.Meta.verbose_name_plural,
                num_row_heading_columns=1)
            data = transpose(data)
        headings = headings[0]

        # get attributes order
        attributes = [model.Meta.attributes[attr_name] for attr_name in model.Meta.attribute_order]

        # sort attributes by header order
        attributes = []
        errors = []
        for verbose_name in headings:
            attr = utils.get_attribute_by_verbose_name(model, verbose_name)
            if attr is None:
                errors.append('Header "{}" does not match any attributes'.format(verbose_name))
            else:
                attributes.append(attr)

        if errors:
            return ([], [], errors, [])

        # read data
        objects = []
        errors = []
        for obj_data in data:
            obj = model()

            for attr, attr_value in zip(attributes, obj_data):
                if not isinstance(attr, RelatedAttribute):
                    value, error = attr.deserialize(attr_value)
                    if error:
                        errors.append(error)
                    else:
                        setattr(obj, attr.name, value)

            objects.append(obj)

        return (attributes, data, errors, objects)

    @classmethod
    def link_model(cls, model, attributes, data, objects, objects_by_primary_attribute):
        """ Read a set of objects from an Excel worksheet

        Args:
            model (:obj:`class`): model
            attributes (:obj:`list` of `Attribute`): attribute order of `data`
            data (:obj:`list` of `list` of `object`): nested list of object data
            objects (:obj:`list`): list of model objects in order of `data`
            objects_by_primary_attribute (:obj:`dict` of `class`: `dict of `object`:`Model`): dictionary of model objects grouped by model

        Returns:
            :obj:`list` of `str`: list of parsing errors
        """

        errors = []
        for obj_data, obj in zip(data, objects):
            for attr, attr_value in zip(attributes, obj_data):
                if isinstance(attr, RelatedAttribute):
                    value, error = attr.deserialize(attr_value, objects_by_primary_attribute)
                    if error:
                        errors.append(error)
                    else:
                        setattr(obj, attr.name, value)

        return errors

    @staticmethod
    def read_sheet(workbook, sheet_name, num_row_heading_columns=0, num_column_heading_rows=0):
        """ Read an Excel sheet in to a two-dimensioanl list

        Args:
            workbook (:obj:`Workbook`): workbook
            sheet_name (:obj:`str`): worksheet name
            num_row_heading_columns (:obj:`int`, optional): number of columns of row headings
            num_column_heading_rows (:obj:`int`, optional): number of rows of column headings

        Returns:
            :obj:`list` of `list`: two-dimensional list of table values
        """
        ws = workbook[sheet_name]

        row_headings = []
        for i_col in range(1, num_row_heading_columns + 1):
            row_headings_col = []
            row_headings.append(row_headings_col)
            for i_row in range(1 + num_column_heading_rows, ws.max_row + 1):
                row_headings_col.append(ws.cell(row=i_row, column=i_col).value)

        column_headings = []
        for i_row in range(1, num_column_heading_rows + 1):
            column_headings_row = []
            column_headings.append(column_headings_row)
            for i_col in range(1 + num_row_heading_columns, ws.max_column + 1):
                column_headings_row.append(ws.cell(row=i_row, column=i_col).value)

        data = []
        for i_row in range(1 + num_column_heading_rows, ws.max_row + 1):
            data_row = []
            data.append(data_row)
            for i_col in range(1 + num_row_heading_columns, ws.max_column + 1):
                data_row.append(ws.cell(row=i_row, column=i_col).value)

        return (data, row_headings, column_headings)
