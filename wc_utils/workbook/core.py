""" Excel utilities

:Author: Jonathan Karr <karr@mssm.edu>
:Date: 2016-11-28
:Copyright: 2016-2018, Karr Lab
:License: MIT
"""

from openpyxl.utils import get_column_letter
import collections


class Workbook(collections.OrderedDict):
    """ Represents an Excel workbook """

    def __eq__(self, other):
        """ Compare two workbooks

        Args:
            other (:obj:`Workbook`): other workbook

        Returns:
            :obj:`bool`: true if workbooks are semantically equal
        """
        if other.__class__ is not self.__class__:
            return False

        if set(self.keys()) != set(other.keys()):
            return False

        for name, sheet in self.items():
            if not sheet == other[name]:
                return False

        return True

    def __ne__(self, other):
        """ Compare two workbooks

        Args:
            other (:obj:`Workbook`): other workbook

        Returns:
            :obj:`bool`: true if workbooks are semantically unequal
        """
        return not self.__eq__(other)

    def difference(self, other):
        """ Get difference with another workbook

        Args:
            other (:obj:`Workbook`): other workbook

        Returns:
            :obj:`dict`: dictionary of differences, grouped by worksheet

        Raises:
            :obj:`ValueError`: if other is not an instance of `workbook`
        """

        if other.__class__ is not self.__class__:
            raise ValueError('`other` must be an instance of `Workbook`')

        diff = WorkbookDifference()
        for name, sheet in self.items():
            if name in other:
                sheet_diff = sheet.difference(other[name])
                if sheet_diff:
                    diff[name] = sheet_diff
            else:
                diff[name] = 'Sheet not in other'

        for name in other.keys():
            if name not in self:
                diff[name] = 'Sheet not in self'

        return diff


class Worksheet(list):
    """ Represents a table of data, such as an Excel worksheet or a csv/tsv file"""

    def __getitem__(self, i_row):
        """ Get a row or a range of rows

        Args:
            i_row (:obj:`int` of :obj:`slice`): row index or range
                of row indices

        Returns:
            :obj:`Row` or :obj:`Worksheet`: row or range of rows
        """
        item = super(Worksheet, self).__getitem__(i_row)
        if isinstance(i_row, slice):
            item = self.__class__(item)
        return item

    def __eq__(self, other):
        """ Compare two worksheets

        Args:
            other (:obj:`Worksheet`): other worksheet

        Returns:
            :obj:`bool`: True if worksheets are semantically equal
        """
        if other.__class__ is not self.__class__:
            return False

        if len(self) != len(other):
            return False

        for row_self, row_other in zip(self, other):
            if not row_self == row_other:
                return False

        return True

    def __ne__(self, other):
        """ Compare two worksheets

        Args:
            other (:obj:`Worksheet`): other worksheet

        Returns:
            :obj:`bool`: True if worksheets are semantically unequal
        """
        return not self.__eq__(other)

    def difference(self, other):
        """ Get difference with another worksheet

        Args:
            other (:obj:`Worksheet`): other worksheet

        Returns:
            :obj:`WorksheeDifference`: dictionary of differences, grouped by row

        Raises:
            :obj:`ValueError`: if other is not an instance of `Worksheet`
        """
        if other.__class__ is not self.__class__:
            raise ValueError('`other` must be an instance of `Worksheet`')

        diff = WorksheetDifference()

        for i_row, row_self in enumerate(self):
            if i_row < len(other):
                diff_row = row_self.difference(other[i_row])
                if diff_row:
                    diff[i_row] = diff_row
            else:
                diff[i_row] = 'Row not in other'

        for i_row in range(len(self), len(other)):
            diff[i_row] = 'Row not in self'

        return diff

    def remove_empty_final_rows(self):
        """ Remove empty final rows """
        for row in reversed(self):
            is_empty = next((False for cell in row if cell not in (None, '')), True)
            if is_empty:
                self.pop()
            else:
                break

    def remove_empty_final_cols(self):
        """ Remove empty final columns """
        max_col = 0
        for row in self:
            for i_rev_col, cell in enumerate(reversed(row)):
                if cell not in (None, ''):
                    max_col = max(max_col, len(row) - i_rev_col)
                    break

        for i_row, row in enumerate(self):
            self[i_row] = row[0:max_col]


class Row(list):
    """ Represents a row in a table of data """

    def __getitem__(self, i_cell):
        """ Get a cell or a range of cells

        Args:
            i_cell (:obj:`int` of :obj:`slice`): cell index or range
                of cell indices

        Returns:
            :obj:`object` or :obj:`Row`: cell or range of cells
        """
        item = super(Row, self).__getitem__(i_cell)
        if isinstance(i_cell, slice):
            item = self.__class__(item)
        return item

    def __eq__(self, other):
        """ Compare rows

        Args:
            other (:obj:`Row`): other row

        Returns:
            :obj:`bool`: True if rows are semantically equal
        """
        if other.__class__ is not self.__class__:
            return False

        if len(self) != len(other):
            return False

        for c_self, c_other in zip(self, other):
            if not (c_self == c_other or (c_self is None and c_other == '') or (c_self == '' and c_other is None)):
                return False

        return True

    def __ne__(self, other):
        """ Compare rows

        Args:
            other (:obj:`Row`): other row

        Returns:
            :obj:`bool`: True if rows are semantically unequal
        """
        return not self.__eq__(other)

    def difference(self, other):
        """ Get difference with another row

        Args:
            other (:obj:`Row`): other row

        Returns:
            :obj:`RowDifference`: dictionary of differences

        Raises:
            :obj:`ValueError`: if other is not an instance of `Row`
        """

        if other.__class__ is not self.__class__:
            raise ValueError('`other` must be an instance of `Row`')

        diff = RowDifference()

        for i_cell, cell_self in enumerate(self):
            if i_cell < len(other):
                diff_cell = self.cell_difference(cell_self, other[i_cell])
                if diff_cell:
                    diff[i_cell] = diff_cell
            else:
                diff[i_cell] = 'Cell not in other'

        for i_cell in range(len(self), len(other)):
            diff[i_cell] = 'Cell not in self'

        return diff

    def cell_difference(self, cell_self, cell_other):
        """ Get difference between cells

        Args:
            cell_self (:obj:`object`): self cell
            cell_other (:obj:`object`): other cell

        Returns:
            :obj:`CellDifference`: difference
        """
        if cell_self == cell_other or (cell_self is None and cell_other == '') or (cell_self == '' and cell_other is None):
            return CellDifference()
        else:
            return CellDifference('{} != {}'.format(cell_self, cell_other))


class WorkbookDifference(dict):
    """ Difference between values of workbook """

    def __str__(self):
        """ Get string representation 

        Returns:
            :obj:`str`: string representation
        """
        diff = ''

        for name, sheet in self.items():
            diff += '\nSheet {}:\n  {}'.format(name, str(sheet).replace('\n', '\n  '))

        return diff[1:]


class WorksheetDifference(collections.OrderedDict):
    """ Difference between values of worksheets """

    def __str__(self):
        """ Get string representation 

        Returns:
            :obj:`str`: string representation
        """
        diff = ''

        for i_row, row in self.items():
            diff += '\nRow {}:\n  {}'.format(i_row + 1, str(row).replace('\n', '\n  '))

        return diff[1:]


class RowDifference(collections.OrderedDict):
    """ Difference between values of rows """

    def __str__(self):
        """ Get string representation 

        Returns:
            :obj:`str`: string representation
        """
        diff = ''

        for i_col, cell_diff in self.items():
            diff += '\nCell {}: {}'.format(get_column_letter(i_col + 1), cell_diff.replace('\n', '\n  '))

        return diff[1:]


class CellDifference(str):
    """ Difference between values of cells """
    pass
