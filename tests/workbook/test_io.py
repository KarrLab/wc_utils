""" Test Excel utilities

:Author: Jonathan Karr <karr@mssm.edu>
:Author: Arthur Goldberg <Arthur.Goldberg@mssm.edu>
:Date: 2016-11-23
:Copyright: 2016-2018, Karr Lab
:License: MIT
"""

from copy import deepcopy
from os import path
from shutil import rmtree
from six import integer_types, string_types
from tempfile import mkdtemp
from wc_utils.workbook import io
from wc_utils.workbook.core import Workbook, Worksheet, Row, Formula
import math
import openpyxl
import unittest


class TestIo(unittest.TestCase):

    def setUp(self):
        # test data set
        wk = self.wk = Workbook()

        ws0 = wk['Ws-0'] = Worksheet()
        ws0.append(Row(['Id', 'Val-1', 'Val-2', 'Val-3']))
        ws0.append(Row(['a0\taa0\naaa0', 1, 2., True]))
        ws0.append(Row([u'b0\u20ac', 3, 4., False]))
        ws0.append(Row(['c0', 5, 6., None]))

        ws1 = wk['Ws-1'] = Worksheet()
        ws1.append(Row(['Id', 'Val-1', 'Val-2']))
        ws1.append(Row(['a1', 1, 2.]))
        ws1.append(Row(['b1', 3, 4.]))
        ws1.append(Row(['c1', 5, 6.]))

        ws2 = wk['Ws-2'] = Worksheet()
        ws2.append(Row(['Id', 'Val-1', 'Val-2']))
        ws2.append(Row(['a2', 1, 2.]))
        ws2.append(Row(['b2', 3, 4.]))
        ws2.append(Row(['c2', 5, 6.]))

        style = self.style = io.WorkbookStyle()
        style['Ws-0'] = io.WorksheetStyle(extra_rows=2, extra_columns=2)
        style['Ws-1'] = io.WorksheetStyle(extra_rows=2, extra_columns=2)
        style['Ws-2'] = io.WorksheetStyle(extra_rows=2, extra_columns=2)

        # create temp directory
        self.tempdir = mkdtemp()

    def tearDown(self):
        # remove temp directory
        rmtree(self.tempdir)

    def test_exceptions_excel(self):
        filename = path.join(self.tempdir, 'test.foo')
        with self.assertRaises(ValueError) as context:
            io.ExcelWriter(filename)
        self.assertIn("Extension of path", str(context.exception))
        self.assertIn("must be '.xlsx'", str(context.exception))

        with self.assertRaises(ValueError) as context:
            io.ExcelReader(filename)
        self.assertIn("Extension of path", str(context.exception))
        self.assertIn("must be '.xlsx'", str(context.exception))

        filename = path.join(self.tempdir, 'test.xlsx')
        wk = deepcopy(self.wk)
        wk['Ws-0'][0][1] = []
        with self.assertRaisesRegex(ValueError, '^Unsupported type '):
            io.ExcelWriter(filename).run(wk, style=self.style)

    def test_read_write_excel(self):
        # write to file
        filename = path.join(self.tempdir, 'test.xlsx')
        io.ExcelWriter(filename).run(self.wk, style=self.style)
        self.assertTrue(path.isfile(filename))

        # write to file with style
        style = self.style
        style['Ws-0'] = io.WorksheetStyle(title_rows=1, head_rows=1, head_columns=1,
                                          head_row_font_bold=True,
                                          head_row_fill_fgcolor='CCCCCC',
                                          row_height=15,
                                          extra_columns=2,
                                          extra_rows=2)
        io.ExcelWriter(filename).run(self.wk, style=style)
        self.assertTrue(path.isfile(filename))

        # read from file
        wk = io.ExcelReader(filename).run()

        # assert content is the same
        ws = wk['Ws-0']
        self.assertIsInstance(ws[1][0], string_types)
        self.assertIsInstance(ws[1][1], integer_types)
        self.assertIsInstance(ws[1][2], integer_types)
        self.assertIsInstance(ws[1][3], bool)
        self.assertEqual(ws[2][0], u'b0\u20ac')
        self.assertEqual(ws[3][3], None)

        self.assertEqual(wk, self.wk)

    def test_write_excel_hidden_rows_cols(self):
        filename = path.join(self.tempdir, 'test.xlsx')
        style = self.style
        style['Ws-0'] = io.WorksheetStyle(head_rows=1, head_columns=1,
                                          head_row_fill_pattern='solid',
                                          head_row_fill_fgcolor='CCCCCC',
                                          row_height=15.01,
                                          col_width=10.,
                                          extra_rows=2, extra_columns=2)
        io.ExcelWriter(filename).run(self.wk, style=style)
        self.assertTrue(path.isfile(filename))

    def test_write_excel_show_inf_rows_cols(self):
        filename = path.join(self.tempdir, 'test.xlsx')
        style = self.style
        style['Ws-0'] = io.WorksheetStyle(head_rows=1, head_columns=1,
                                          head_row_fill_pattern='solid',
                                          head_row_fill_fgcolor='CCCCCC',
                                          row_height=float('nan'),
                                          col_width=10.,
                                          extra_rows=float('inf'), extra_columns=float('inf'))
        io.ExcelWriter(filename).run(self.wk, style=style)
        self.assertTrue(path.isfile(filename))

    def test_write_excel_no_data(self):
        filename = path.join(self.tempdir, 'test.xlsx')
        style = self.style
        style['Ws-0'] = io.WorksheetStyle(head_rows=0, head_columns=0,
                                          head_row_fill_pattern='solid',
                                          head_row_fill_fgcolor='CCCCCC',
                                          row_height=float('nan'),
                                          col_width=10.,
                                          extra_rows=float('inf'), extra_columns=float('inf'))
        self.wk['Ws-0'] = Workbook()
        io.ExcelWriter(filename).run(self.wk, style=style)
        self.assertTrue(path.isfile(filename))

    def test_write_excel_pattern_error(self):
        filename = path.join(self.tempdir, 'test.xlsx')
        style = self.style
        style['Ws-0'] = io.WorksheetStyle(head_rows=0, head_columns=0,
                                          head_row_fill_pattern='UNDEFINED_PATTERN',
                                          head_row_fill_fgcolor='CCCCCC',
                                          row_height=float('nan'),
                                          col_width=10.,
                                          extra_rows=float('inf'), extra_columns=float('inf'))
        with self.assertRaisesRegex(ValueError, 'Unsupported pattern'):
            io.ExcelWriter(filename).run(self.wk, style=style)

        filename = path.join(self.tempdir, 'test.xlsx')
        style = self.style
        style['Ws-0'] = io.WorksheetStyle(head_rows=0, head_columns=0,
                                          title_row_fill_pattern='UNDEFINED_PATTERN',
                                          title_row_fill_fgcolor='CCCCCC',
                                          row_height=float('nan'),
                                          col_width=10.,
                                          extra_rows=float('inf'), extra_columns=float('inf'))
        with self.assertRaisesRegex(ValueError, 'Unsupported pattern'):
            io.ExcelWriter(filename).run(self.wk, style=style)

    def test_write_excel_row_validation(self):
        filename = path.join(self.tempdir, 'test.xlsx')
        style = self.style
        style['Ws-0'] = io.WorksheetStyle(head_rows=1, head_columns=0,
                                          head_row_fill_pattern='solid',
                                          head_row_fill_fgcolor='CCCCCC',
                                          row_height=15.01,
                                          col_width=10.,
                                          extra_rows=2, extra_columns=2)
        validation = io.WorkbookValidation()
        validation['Ws-0'] = io.WorksheetValidation(fields=[
            io.FieldValidation(input_title='Enter a identifier', input_message='A unique string',
                               type=io.FieldValidationType.length,
                               criterion=io.FieldValidationCriterion['<='],
                               allowed_scalar_value=255),
            None,
            io.FieldValidation(input_title='Enter a second value', input_message='A float',
                               type=io.FieldValidationType.decimal, criterion=io.FieldValidationCriterion['>='],
                               minimum_scalar_value=-1000.),
            io.FieldValidation(input_title='Enter a third value', input_message='A float',
                               type=io.FieldValidationType.any),
        ])

        io.ExcelWriter(filename).run(self.wk, style=style, validation=validation)
        self.assertTrue(path.isfile(filename))

    def test_write_excel_col_validation(self):
        filename = path.join(self.tempdir, 'test.xlsx')
        style = self.style
        style['Ws-0'] = io.WorksheetStyle(head_rows=0, head_columns=1,
                                          head_row_fill_pattern='solid',
                                          head_row_fill_fgcolor='CCCCCC',
                                          row_height=15.01,
                                          col_width=10.,
                                          extra_rows=2, extra_columns=2)
        validation = io.WorkbookValidation()
        validation['Ws-0'] = io.WorksheetValidation(orientation=io.WorksheetValidationOrientation.column, fields=[
            None,
            io.FieldValidation(input_title='Enter a identifier', input_message='A unique string',
                               type=io.FieldValidationType.length,
                               criterion=io.FieldValidationCriterion['<='],
                               allowed_scalar_value=255,),
            io.FieldValidation(input_title='Enter a second value', input_message='A float',
                               type=io.FieldValidationType.decimal, criterion=io.FieldValidationCriterion['>='],
                               minimum_scalar_value=-1000.),
            io.FieldValidation(input_title='Enter a third value', input_message='A float',
                               type=io.FieldValidationType.any),
        ])

        io.ExcelWriter(filename).run(self.wk, style=style, validation=validation)
        self.assertTrue(path.isfile(filename))

    def test_write_excel_long_message_truncation(self):
        filename = path.join(self.tempdir, 'test.xlsx')
        style = self.style
        style['Ws-0'] = io.WorksheetStyle(head_rows=1, head_columns=0,
                                          head_row_fill_pattern='solid',
                                          head_row_fill_fgcolor='CCCCCC',
                                          row_height=15.01,
                                          col_width=10.,
                                          extra_rows=2, extra_columns=2)
        validation = io.WorkbookValidation()
        validation['Ws-0'] = io.WorksheetValidation(fields=[
            io.FieldValidation(input_title='Enter a identifier' * 100, input_message='A unique string' * 100,
                               error_title='Enter a identifier' * 100, error_message='A unique string' * 100,
                               type=io.FieldValidationType.length,
                               criterion=io.FieldValidationCriterion['<='],
                               allowed_scalar_value=255),
            None,
            io.FieldValidation(input_title='Enter a second value', input_message='A float',
                               type=io.FieldValidationType.decimal, criterion=io.FieldValidationCriterion['>='],
                               minimum_scalar_value=-1000.),
            io.FieldValidation(input_title='Enter a third value', input_message='A float',
                               type=io.FieldValidationType.any),
        ])

        io.ExcelWriter(filename).run(self.wk, style=style, validation=validation)
        self.assertTrue(path.isfile(filename))

    def test_FieldValidation_get_options(self):
        fv = io.FieldValidation(input_title='input_title', input_message='input_message', show_input=True,
                                type=io.FieldValidationType.any, criterion=io.FieldValidationCriterion[
                                    '<='], allowed_scalar_value='allowed_scalar_value',
                                minimum_scalar_value=-2., maximum_scalar_value=2., allowed_list_values=['a', 'b', 'c'],
                                show_dropdown=True, ignore_blank=True,
                                error_type=io.FieldValidationErrorType.stop, error_title='error_title', error_message='error_message',
                                show_error=True)
        self.assertEqual(fv.get_options(), {
            'input_title': 'input_title',
            'input_message': 'input_message',
            'show_input': True,
            'validate': 'any',
            'criteria': '<=',
            'value': 'allowed_scalar_value',
            'minimum': -2.,
            'maximum': 2.,
            'source': ['a', 'b', 'c'],
            'dropdown': True,
            'ignore_blank': True,
            'error_type': 'stop',
            'error_title': 'error_title',
            'error_message': 'error_message',
            'show_error': True,
        })

    def test_excel_read_valid_types(self):
        wb = openpyxl.Workbook()
        ws = wb.create_sheet('Sheet-1')

        cell = ws.cell(row=1, column=1)
        cell.data_type = openpyxl.cell.cell.TYPE_STRING
        cell.value = 'A1'

        cell = ws.cell(row=2, column=1)
        cell.data_type = openpyxl.cell.cell.TYPE_NUMERIC
        cell.value = 2.5

        cell = ws.cell(row=3, column=1)
        cell.data_type = openpyxl.cell.cell.TYPE_NUMERIC
        cell.value = None

        cell = ws.cell(row=4, column=1)
        cell.data_type = openpyxl.cell.cell.TYPE_BOOL
        cell.value = True

        cell = ws.cell(row=5, column=1)
        cell.data_type = openpyxl.cell.cell.TYPE_BOOL
        cell.value = False

        cell = ws.cell(row=6, column=1)
        cell.data_type = openpyxl.cell.cell.TYPE_NULL
        cell.value = None

        cell = ws.cell(row=7, column=1)
        cell.data_type = openpyxl.cell.cell.TYPE_INLINE
        cell.value = '<b>A7</b>'

        filename = path.join(self.tempdir, 'test.xlsx')
        wb.save(filename)

        wb2 = io.ExcelReader(filename).run()
        self.assertEqual(wb2['Sheet-1'][0][0], 'A1')
        self.assertEqual(wb2['Sheet-1'][1][0], 2.5)
        self.assertEqual(wb2['Sheet-1'][2][0], None)
        self.assertEqual(wb2['Sheet-1'][3][0], True)
        self.assertEqual(wb2['Sheet-1'][4][0], False)
        self.assertEqual(wb2['Sheet-1'][5][0], None)
        self.assertEqual(wb2['Sheet-1'][6][0], '<b>A7</b>')

    def test_excel_read_valid_types_empty(self):
        wb = openpyxl.Workbook()
        ws = wb.create_sheet('Sheet-1')

        cell = ws.cell(row=1, column=1)
        cell.set_explicit_value(None, openpyxl.cell.cell.TYPE_STRING)

        cell = ws.cell(row=1, column=2)
        cell.set_explicit_value('', openpyxl.cell.cell.TYPE_STRING)

        cell = ws.cell(row=2, column=1)
        cell.set_explicit_value(None, openpyxl.cell.cell.TYPE_INLINE)

        cell = ws.cell(row=2, column=2)
        cell.set_explicit_value('', openpyxl.cell.cell.TYPE_INLINE)

        cell = ws.cell(row=3, column=1)
        cell.set_explicit_value(None, openpyxl.cell.cell.TYPE_NUMERIC)

        cell = ws.cell(row=3, column=2)
        cell.set_explicit_value('', openpyxl.cell.cell.TYPE_NUMERIC)

        cell = ws.cell(row=4, column=1)
        cell.set_explicit_value(None, openpyxl.cell.cell.TYPE_BOOL)

        cell = ws.cell(row=4, column=2)
        cell.set_explicit_value('', openpyxl.cell.cell.TYPE_BOOL)

        cell = ws.cell(row=5, column=1)
        cell.set_explicit_value(None, openpyxl.cell.cell.TYPE_NULL)

        cell = ws.cell(row=5, column=2)
        cell.set_explicit_value('', openpyxl.cell.cell.TYPE_NULL)

        cell = ws.cell(row=6, column=2)
        cell.set_explicit_value('end', openpyxl.cell.cell.TYPE_STRING)  # to force max row and column

        filename = path.join(self.tempdir, 'test.xlsx')
        wb.save(filename)

        wb2 = openpyxl.load_workbook(filename)
        self.assertEqual(wb2['Sheet-1'].cell(row=1, column=1).data_type, openpyxl.cell.cell.TYPE_NULL)
        self.assertEqual(wb2['Sheet-1'].cell(row=1, column=2).data_type, openpyxl.cell.cell.TYPE_INLINE)
        self.assertEqual(wb2['Sheet-1'].cell(row=2, column=1).data_type, openpyxl.cell.cell.TYPE_NULL)
        self.assertEqual(wb2['Sheet-1'].cell(row=2, column=2).data_type, openpyxl.cell.cell.TYPE_INLINE)
        self.assertEqual(wb2['Sheet-1'].cell(row=3, column=1).data_type, openpyxl.cell.cell.TYPE_NULL)
        self.assertEqual(wb2['Sheet-1'].cell(row=3, column=2).data_type, openpyxl.cell.cell.TYPE_NULL)
        self.assertEqual(wb2['Sheet-1'].cell(row=4, column=1).data_type, openpyxl.cell.cell.TYPE_NULL)
        self.assertEqual(wb2['Sheet-1'].cell(row=4, column=2).data_type, openpyxl.cell.cell.TYPE_BOOL)
        self.assertEqual(wb2['Sheet-1'].cell(row=5, column=1).data_type, openpyxl.cell.cell.TYPE_NULL)
        self.assertEqual(wb2['Sheet-1'].cell(row=5, column=2).data_type, openpyxl.cell.cell.TYPE_NULL)

        wb2 = io.ExcelReader(filename).run()
        self.assertEqual(wb2['Sheet-1'][0][0], None)
        self.assertEqual(wb2['Sheet-1'][0][1], None)
        self.assertEqual(wb2['Sheet-1'][1][0], None)
        self.assertEqual(wb2['Sheet-1'][1][1], None)

        self.assertEqual(wb2['Sheet-1'][2][0], None)
        self.assertEqual(wb2['Sheet-1'][2][1], None)
        self.assertEqual(wb2['Sheet-1'][3][0], None)
        self.assertEqual(wb2['Sheet-1'][3][1], None)

        self.assertEqual(wb2['Sheet-1'][4][0], None)
        self.assertEqual(wb2['Sheet-1'][4][1], None)

    def test_excel_read_formula_boolean(self):
        wb = openpyxl.Workbook()
        ws = wb.create_sheet('Sheet-1')

        cell = ws.cell(row=1, column=1)
        cell.data_type = openpyxl.cell.cell.TYPE_FORMULA
        cell.value = '=TRUE()'

        cell = ws.cell(row=1, column=2)
        cell.data_type = openpyxl.cell.cell.TYPE_FORMULA
        cell.value = '=TRUE'

        cell = ws.cell(row=2, column=1)
        cell.data_type = openpyxl.cell.cell.TYPE_FORMULA
        cell.value = '=FALSE()'

        cell = ws.cell(row=2, column=2)
        cell.data_type = openpyxl.cell.cell.TYPE_FORMULA
        cell.value = '=FALSE'

        filename = path.join(self.tempdir, 'test.xlsx')
        wb.save(filename)

        wb2 = io.ExcelReader(filename).run()
        self.assertEqual(wb2['Sheet-1'][0][0], True)
        self.assertEqual(wb2['Sheet-1'][0][1], True)
        self.assertEqual(wb2['Sheet-1'][1][0], False)
        self.assertEqual(wb2['Sheet-1'][1][1], False)

    def test_excel_read_formula_algebra(self):
        wb = openpyxl.Workbook()
        ws = wb.create_sheet('Sheet-1')

        cell = ws.cell(row=1, column=1)
        cell.data_type = openpyxl.cell.cell.TYPE_FORMULA
        cell.value = '=1+2'

        filename = path.join(self.tempdir, 'test.xlsx')
        wb.save(filename)

        io.ExcelReader(filename).run()

    def test_excel_read_formula_cache_string(self):
        wb = openpyxl.Workbook()
        ws = wb.create_sheet('Sheet-1')

        cell = ws.cell(row=1, column=1)
        cell.data_type = openpyxl.cell.cell.TYPE_FORMULA_CACHE_STRING
        cell.value = '=1+2'

        filename = path.join(self.tempdir, 'test.xlsx')
        wb.save(filename)

        io.ExcelReader(filename).run()

    def test_excel_read_error(self):
        wb = openpyxl.Workbook()
        ws = wb.create_sheet('Sheet-1')

        cell = ws.cell(row=1, column=1)
        cell.data_type = openpyxl.cell.cell.TYPE_ERROR
        cell.value = '#NAME?'

        filename = path.join(self.tempdir, 'test.xlsx')
        wb.save(filename)

        with self.assertRaisesRegex(ValueError, 'Errors are not supported'):
            io.ExcelReader(filename).run()

    def test_excel_ignore_empty_final_rows_and_cols(self):
        wb = openpyxl.Workbook()
        ws = wb.create_sheet('Sheet-1')

        cell = ws.cell(row=1, column=1)
        cell.data_type = openpyxl.cell.cell.TYPE_STRING
        cell.value = 'A1'

        cell = ws.cell(row=1, column=2)
        cell.data_type = openpyxl.cell.cell.TYPE_STRING
        cell.value = 'A2'

        cell = ws.cell(row=1, column=3)
        cell.data_type = openpyxl.cell.cell.TYPE_NULL
        cell.value = None

        cell = ws.cell(row=2, column=1)
        cell.data_type = openpyxl.cell.cell.TYPE_STRING
        cell.value = 'B1'

        cell = ws.cell(row=2, column=2)
        cell.data_type = openpyxl.cell.cell.TYPE_STRING
        cell.value = 'B2'

        cell = ws.cell(row=2, column=3)
        cell.data_type = openpyxl.cell.cell.TYPE_NULL
        cell.value = None

        cell = ws.cell(row=3, column=1)
        cell.data_type = openpyxl.cell.cell.TYPE_NULL
        cell.value = None

        cell = ws.cell(row=3, column=2)
        cell.data_type = openpyxl.cell.cell.TYPE_NULL
        cell.value = None

        cell = ws.cell(row=3, column=3)
        cell.data_type = openpyxl.cell.cell.TYPE_NULL
        cell.value = None

        filename = path.join(self.tempdir, 'test.xlsx')
        wb.save(filename)

        wb = io.ExcelReader(filename).run()

        self.assertEqual(len(wb['Sheet-1']), 2)
        self.assertEqual(list(wb['Sheet-1'][0]), ['A1', 'A2'])
        self.assertEqual(list(wb['Sheet-1'][1]), ['B1', 'B2'])

    def test_excel_merge_cells(self):
        wb = Workbook()

        ws0 = wb['Ws-0'] = Worksheet()
        ws0.append(Row([None, 'Vals', 'Vals', 'Vals']))
        ws0.append(Row([None, 'Vals 1-2', 'Vals 1-2', None]))
        ws0.append(Row([None, 'Vals 1-2', 'Vals 1-2', None]))
        ws0.append(Row(['Id', 'Val-1', 'Val-2', 'Val-3']))
        ws0.append(Row(['a0\taa0\naaa0', 1, 2., True]))
        ws0.append(Row([u'b0\u20ac', 3, 4., False]))
        ws0.append(Row(['c0', 5, 6., None]))

        style = io.WorkbookStyle()
        style['Ws-0'] = io.WorksheetStyle(head_rows=4,
                                          merge_ranges=[(0, 0, 0, 0), (0, 1, 0, 3), (1, 1, 2, 2)],
                                          blank_head_fill_fgcolor='EEEEEE',
                                          merged_head_fill_fgcolor='AAAAAA')

        filename = path.join(self.tempdir, 'test.xlsx')
        io.ExcelWriter(filename).run(wb, style=style)
        wb_2 = io.ExcelReader(filename).run()
        self.assertEqual(wb_2, wb)

        filename = path.join(self.tempdir, 'test-*.csv')
        io.SeparatedValuesWriter(filename).run(wb, style=style)
        wb_2 = io.SeparatedValuesReader(filename).run()
        self.assertEqual(wb_2, wb)

    def test_excel_merge_cells_error(self):
        wb = Workbook()

        ws0 = wb['Ws-0'] = Worksheet()
        ws0.append(Row([None, 'Vals', None, None]))
        ws0.append(Row([None, 'Vals 1-2', None, None]))
        ws0.append(Row([None, None, None, None]))
        ws0.append(Row(['Id', 'Val-1', 'Val-2', 'Val-3']))
        ws0.append(Row(['a0\taa0\naaa0', 1, 2., True]))
        ws0.append(Row([u'b0\u20ac', 3, 4., False]))
        ws0.append(Row(['c0', 5, 6., None]))

        style = io.WorkbookStyle()
        style['Ws-0'] = io.WorksheetStyle(merge_ranges=[(0, 0, 0, 0), (0, 1, 0, 3), (1, 1, 2, 2), (3, 0, 3, 3)])

        filename = path.join(self.tempdir, 'test.xlsx')
        with self.assertRaisesRegex(ValueError, 'can have at most 1 value'):
            io.ExcelWriter(filename).run(wb, style=style)

    def test_formula_hyperlink(self):
        wb = Workbook()
        ws0 = wb['Ws'] = Worksheet()
        ws0.append(Row(['abc', 'def', 'ghi']))
        ws0[0][0] = Formula('="abc"', 'abc')

        wb_1 = Workbook()
        ws0 = wb_1['Ws'] = Worksheet()
        ws0.append(Row(['abc', 'def', 'ghi']))
        ws0[0][0] = Formula('="abc"', 'abc')

        style = io.WorkbookStyle()        
        style['Ws'] = io.WorksheetStyle(
          hyperlinks=[io.Hyperlink(0, 1, 'https://google.com', tip='Click to view def')])

        filename = path.join(self.tempdir, 'test.xlsx')
        io.ExcelWriter(filename).run(wb, style=style)
        wb_2 = io.ExcelReader(filename).run()
        wb_2['Ws'][0][0].value = 'abc'
        self.assertEqual(wb_2, wb_1)

        filename = path.join(self.tempdir, 'test*.csv')
        io.SeparatedValuesWriter(filename).run(wb, style=style)
        wb_2 = io.SeparatedValuesReader(filename).run()
        wb_2['Ws'][0][0] = Formula('="abc"', wb_2['Ws'][0][0])
        self.assertEqual(wb_2, wb_1)

    def test_exceptions_csv(self):
        for method in [io.SeparatedValuesWriter, io.SeparatedValuesReader]:
            filename = path.join(self.tempdir, 'test.foo')
            with self.assertRaises(ValueError) as context:
                method(filename)
            self.assertIn("Extension of path", str(context.exception))
            self.assertIn("must be one of '.csv' or '.tsv'", str(context.exception))

            filename = path.join(self.tempdir, '*', 'test.csv')
            with self.assertRaises(ValueError) as context:
                method(filename)
            self.assertIn("cannot have glob pattern '*' in its directory name", str(context.exception))

            filename = path.join(self.tempdir, 'test**.csv')
            with self.assertRaises(ValueError) as context:
                method(filename)
            self.assertIn("cannot have multiple glob patterns '*' in its base name", str(context.exception))

    def test_read_write_csv(self):
        # write to files
        filename_pattern = path.join(self.tempdir, 'test-*.csv')
        io.SeparatedValuesWriter(filename_pattern).run(self.wk)
        self.assertTrue(path.isfile(filename_pattern.replace('*', '{}').format('Ws-0')))
        self.assertTrue(path.isfile(filename_pattern.replace('*', '{}').format('Ws-1')))
        self.assertTrue(path.isfile(filename_pattern.replace('*', '{}').format('Ws-2')))

        # read from files
        wk = io.SeparatedValuesReader(filename_pattern).run()

        # assert content is the same
        ws = wk['Ws-0']
        self.assertIsInstance(ws[1][0], string_types)
        self.assertIsInstance(ws[1][1], integer_types)
        self.assertIsInstance(ws[1][2], float)
        self.assertIsInstance(ws[1][3], bool)
        self.assertEqual(ws[2][0], u'b0\u20ac')
        self.assertEqual(ws[3][3], None)

        self.assertEqual(wk, self.wk)

    def test_read_write_tsv(self):
        # write to files
        filename_pattern = path.join(self.tempdir, 'test-*.tsv')
        io.SeparatedValuesWriter(filename_pattern).run(self.wk)
        self.assertTrue(path.isfile(filename_pattern.replace('*', '{}').format('Ws-0')))
        self.assertTrue(path.isfile(filename_pattern.replace('*', '{}').format('Ws-1')))
        self.assertTrue(path.isfile(filename_pattern.replace('*', '{}').format('Ws-2')))

        # read from files
        wk = io.SeparatedValuesReader(filename_pattern).run()

        # assert content is the same
        ws = wk['Ws-0']
        self.assertIsInstance(ws[1][0], string_types)
        self.assertIsInstance(ws[1][1], integer_types)
        self.assertIsInstance(ws[1][2], float)
        self.assertIsInstance(ws[1][3], bool)
        self.assertEqual(ws[2][0], u'b0\u20ac')
        self.assertEqual(ws[3][3], None)

        self.assertEqual(wk, self.wk)

    def test_read_write_csv_no_glob(self):
        wb = Workbook()
        ws = wb['Sheet1'] = Worksheet()
        ws.append(Row(['a', 'b', 'c']))
        ws.append(Row(['d', 'e', 'f']))
        ws.append(Row(['g', 'h', 'i']))

        filename = path.join(self.tempdir, 'test.csv')
        io.SeparatedValuesWriter(filename).run(wb)

        wb2 = io.SeparatedValuesReader(filename).run()
        
        wb2['Sheet1'] = wb2.pop('')
        self.assertEqual(wb2, wb)

        filename2 = path.join(self.tempdir, 'test2-*.csv')
        io.convert(filename, filename2)
        wb2 = io.SeparatedValuesReader(filename2).run()
        wb2['Sheet1'] = wb2.pop('')
        self.assertEqual(wb2, wb)

        filename3 = path.join(self.tempdir, 'test3.csv')
        io.convert(filename, filename3)
        wb2 = io.SeparatedValuesReader(filename3).run()
        wb2['Sheet1'] = wb2.pop('')
        self.assertEqual(wb2, wb)

    def test_read_write_csv_no_glob_error(self):
        wb = Workbook()
        ws = wb['Sheet1'] = Worksheet()
        ws.append(Row(['a', 'b', 'c']))
        ws.append(Row(['d', 'e', 'f']))
        ws = wb['Sheet2'] = Worksheet()
        ws.append(Row(['g', 'h', 'i']))
        ws.append(Row(['j', 'k', 'l']))

        filename = path.join(self.tempdir, 'test.csv')
        with self.assertRaisesRegex(ValueError, 'must have a glob pattern'):
          io.SeparatedValuesWriter(filename).run(wb)

    def test_write_read(self):
        file = path.join(self.tempdir, 'test.xlsx')
        io.write(file, self.wk, style=self.style)
        wk = io.read(file)
        self.assertEqual(wk, self.wk)

        file = path.join(self.tempdir, 'test-*.csv')
        io.write(file, self.wk)
        wk = io.read(file)
        self.assertEqual(wk, self.wk)

    def test_convert(self):
        source = path.join(self.tempdir, 'test.xlsx')
        io.ExcelWriter(source).run(self.wk, style=self.style)

        # copy excel->sv
        dest = path.join(self.tempdir, 'test-*.csv')
        io.convert(source, dest)
        wk = io.SeparatedValuesReader(dest).run()
        self.assertEqual(wk, self.wk)

        # copy sv->excel
        source = path.join(self.tempdir, 'test-*.csv')
        dest = path.join(self.tempdir, 'test2.xlsx')
        io.convert(source, dest, style=self.style)
        wk = io.ExcelReader(dest).run()
        self.assertEqual(wk, self.wk)

        # copy same format - excel
        source = path.join(self.tempdir, 'test.xlsx')
        dest = path.join(self.tempdir, 'test3.xlsx')
        io.convert(source, dest, style=self.style)
        wk = io.ExcelReader(dest).run()
        self.assertEqual(wk, self.wk)

        # copy same format - csv
        source = path.join(self.tempdir, 'test-*.csv')
        dest = path.join(self.tempdir, 'test2-*.csv')
        io.convert(source, dest)
        wk = io.SeparatedValuesReader(dest).run()
        self.assertEqual(wk, self.wk)

        # negative examples
        source = path.join(self.tempdir, 'test.xlsx')
        dest = path.join(self.tempdir, 'test.xlsx')
        self.assertRaises(ValueError, lambda: io.convert(source, dest))

        source = path.join(self.tempdir, 'test.xlsx')
        dest = path.join(self.tempdir, 'test.xlsx2')
        self.assertRaises(ValueError, lambda: io.convert(source, dest))

        source = path.join(self.tempdir, 'test.xlsx2')
        dest = path.join(self.tempdir, 'test.xlsx')
        self.assertRaises(ValueError, lambda: io.convert(source, dest))

    def test_convert_excel_to_csv(self):
        filename_excel = path.join(self.tempdir, 'test.xlsx')
        io.ExcelWriter(filename_excel).run(self.wk, style=self.style)

        filename_pattern_separated_values = path.join(self.tempdir, 'test-*.csv')
        io.convert(filename_excel, filename_pattern_separated_values)
        self.assertTrue(path.isfile(filename_pattern_separated_values.replace('*', '{}').format('Ws-0')))
        self.assertTrue(path.isfile(filename_pattern_separated_values.replace('*', '{}').format('Ws-1')))
        self.assertTrue(path.isfile(filename_pattern_separated_values.replace('*', '{}').format('Ws-2')))

        # read from files
        wk = io.SeparatedValuesReader(filename_pattern_separated_values).run()

        # assert content is the same
        self.assertEqual(wk, self.wk)

    def test_convert_csv_to_excel(self):
        filename_pattern_separated_values = path.join(self.tempdir, 'test-*.csv')
        io.SeparatedValuesWriter(filename_pattern_separated_values).run(self.wk)

        filename_excel = path.join(self.tempdir, 'test.xlsx')
        io.convert(filename_pattern_separated_values, filename_excel,
                   style=self.style)
        self.assertTrue(path.isfile(filename_excel))

        # read from files
        wk = io.ExcelReader(filename_excel).run()

        # assert content is the same
        self.assertEqual(wk, self.wk)

    def test_convert_with_worksheet_order(self):
        source = path.join(self.tempdir, 'test.xlsx')
        io.ExcelWriter(source).run(self.wk, style=self.style)

        dest = path.join(self.tempdir, 'test-2.xlsx')

        io.convert(source, dest, worksheet_order=['Ws-3'], ignore_extra_sheets=True,
                   style=self.style
                   )
        wk = io.ExcelReader(dest).run()
        self.assertEqual(set(wk.keys()), set(['Ws-0', 'Ws-1', 'Ws-2']))

        with self.assertRaisesRegex(ValueError, ' missing worksheets:'):
            io.convert(source, dest, worksheet_order=['Ws-3'], ignore_extra_sheets=False)

    def test_convert_exceptions(self):
        source = path.join(self.tempdir, 'test.xlsx')
        io.ExcelWriter(source).run(self.wk, style=self.style)

        # copy excel->sv
        dest = path.join(self.tempdir, 'test-*.csv')
        io.convert(source, dest)
        wk = io.SeparatedValuesReader(dest).run()
        self.assertEqual(wk, self.wk)

        # copy sv->excel
        source = path.join(self.tempdir, 'test-2-*.csv')
        dest = path.join(self.tempdir, 'test2-*.csv')
        with self.assertRaisesRegex(ValueError, ' does not match any files$'):
            io.convert(source, dest)

    def test_get_reader(self):
        self.assertEqual(io.get_reader('.xlsx'), io.ExcelReader)
        self.assertEqual(io.get_reader('.csv'), io.SeparatedValuesReader)
        self.assertEqual(io.get_reader('.tsv'), io.SeparatedValuesReader)
        self.assertRaises(ValueError, lambda: io.get_reader('.txt'))

    def test_get_writer(self):
        self.assertEqual(io.get_writer('.xlsx'), io.ExcelWriter)
        self.assertEqual(io.get_writer('.csv'), io.SeparatedValuesWriter)
        self.assertEqual(io.get_writer('.tsv'), io.SeparatedValuesWriter)
        self.assertRaises(ValueError, lambda: io.get_writer('.txt'))

    def test_get_sheet_names(self):
        filename_pattern = path.join(self.tempdir, 'test-*.csv')
        io.SeparatedValuesWriter(filename_pattern).run(self.wk)
        self.assertTrue(path.isfile(filename_pattern.replace('*', '{}').format('Ws-0')))
        self.assertTrue(path.isfile(filename_pattern.replace('*', '{}').format('Ws-1')))
        self.assertTrue(path.isfile(filename_pattern.replace('*', '{}').format('Ws-2')))

        # read from files
        filename_pattern = path.join(self.tempdir, 'test-2-*.csv')
        with self.assertRaisesRegex(ValueError, 'does not match any files$'):
            io.SeparatedValuesReader(filename_pattern).run()

    def test_read_empty_worksheet(self):
        wk = Workbook()
        ws = wk['Ws'] = Worksheet()

        filename = path.join(self.tempdir, 'test.xlsx')
        io.ExcelWriter(filename).run(wk)
        wk2 = io.ExcelReader(filename).run()
        self.assertEqual(list(wk2.keys()), ['Ws'])
        self.assertEqual(wk2['Ws'], Worksheet())

        filename = path.join(self.tempdir, 'test-*.csv')
        io.SeparatedValuesWriter(filename).run(wk)
        wk2 = io.SeparatedValuesReader(filename).run()
        self.assertEqual(list(wk2.keys()), ['Ws'])
        self.assertEqual(wk2['Ws'], Worksheet())
