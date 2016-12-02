""" Excel utilities

:Author: Jonathan Karr <karr@mssm.edu>
:Date: 2016-11-28
:Copyright: 2016, Karr Lab
:License: MIT
"""

from collections import OrderedDict
from glob import glob
from math import isnan
from openpyxl import Workbook as XlsWorkbook, load_workbook
from openpyxl.cell.cell import Cell as CellType
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.colors import Color
import pyexcel
import six


def read_excel(filename):
    """ Read data from Excel workbook

    Args:
        filename (:obj:`str`): path to Excel file

    Returns:
        :obj:`Workbook`: python representation of data
    """
    workbook = Workbook()
    xls_workbook = load_workbook(filename=filename)
    for sheet_name in xls_workbook.get_sheet_names():
        xls_worksheet = xls_workbook[sheet_name]
        worksheet = workbook.worksheets[sheet_name] = Worksheet()

        for i_row in range(1, xls_worksheet.max_row + 1):
            row = Row()
            worksheet.rows.append(row)
            for i_col in range(1, xls_worksheet.max_column + 1):
                cell = Cell(xls_worksheet.cell(row=i_row, column=i_col).value)
                row.cells.append(cell)

    return workbook


def write_excel(workbook, filename, style=None):
    """ Read data to an Excel workbook

    Args:
        workbook (:obj:`Workbook`): python representation of data
        filename (:obj:`str`): path to Excel file
        style (:obj:`WorkbookStyle`): workbook style
    """

    style = style or WorkbookStyle()

    xls_workbook = XlsWorkbook()
    xls_workbook.remove_sheet(xls_workbook.active)

    for sheet_name, worksheet in workbook.worksheets.items():
        xls_worksheet = xls_workbook.create_sheet(sheet_name)

        alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        if sheet_name in style.worksheets:
            sheet_style = style.worksheets[sheet_name]
        else:
            sheet_style = WorksheetStyle()

        frozen_rows = sheet_style.head_rows
        frozen_columns = sheet_style.head_columns
        row_height = sheet_style.row_height
        head_font = Font(bold=sheet_style.head_row_font_bold)
        kwargs = {}
        if sheet_style.head_row_fill_pattern:
            kwargs['patternType'] = sheet_style.head_row_fill_pattern
        if sheet_style.head_row_fill_fgcolor:
            kwargs['fgColor'] = sheet_style.head_row_fill_fgcolor
        head_fill = PatternFill(**kwargs)

        for i_row, row in enumerate(worksheet.rows):
            for i_col, cell in enumerate(row.cells):
                xls_cell = xls_worksheet.cell(row=i_row + 1, column=i_col + 1)

                value = cell.value
                if value is None:
                    pass
                elif isinstance(value, six.string_types):
                    data_type = CellType.TYPE_STRING
                elif isinstance(value, bool):
                    data_type = CellType.TYPE_BOOL
                elif isinstance(value, six.integer_types):
                    value = float(value)
                    data_type = CellType.TYPE_NUMERIC
                elif isinstance(value, float):
                    data_type = CellType.TYPE_NUMERIC
                else:
                    raise ValueError('Unsupported type {}'.format(value.__class__.__name__))

                if value is not None:
                    xls_cell.set_explicit_value(value=value, data_type=data_type)

                xls_cell.alignment = alignment

                if i_row < frozen_rows:
                    if head_font:
                        xls_cell.font = head_font
                    if head_fill:
                        xls_cell.fill = head_fill

            if not isnan(row_height):
                xls_worksheet.row_dimensions[i_row + 1].height = row_height

        xls_worksheet.freeze_panes = xls_worksheet.cell(row=frozen_rows + 1, column=frozen_columns + 1)

    xls_workbook.save(filename)


def read_separated_values(filename_pattern):
    """ Read data from a set of [tc]sv files

    Args:
        filename_pattern (:obj:`str`): pattern for file paths e.g. 'workbook-*.csv'

    Returns:
        :obj:`Workbook`: python representation of data
    """

    if filename_pattern[-4:] not in ('.csv', '.tsv'):
        raise ValueError('Extension of `filename_pattern` must match be one of "csv" or "tsv"')

    if filename_pattern.count('*') != 1:
        raise ValueError('`filename_pattern` must have one glob pattern "*"')

    workbook = Workbook()
    i_glob = filename_pattern.find('*')
    for filename in glob(filename_pattern):
        sheet_name = filename[i_glob:i_glob + len(filename) - len(filename_pattern) + 1]
        worksheet = workbook.worksheets[sheet_name] = Worksheet()
        sv_worksheet = pyexcel.get_sheet(file_name=filename)

        for sv_row in sv_worksheet.row:
            row = Row()
            worksheet.rows.append(row)
            for sv_cell in sv_row:
                if sv_cell == '':
                    sv_cell = None
                elif sv_cell == 'True':
                    sv_cell = True
                elif sv_cell == 'False':
                    sv_cell = False

                row.cells.append(Cell(sv_cell))

    return workbook


def write_separated_values(workbook, filename_pattern):
    """ Write data to a set of [tc]sv files

    Args:
        workbook (:obj:`Workbook`): python representation of data
        filename_pattern (:obj:`str`): template for file paths, e.g. 'workbook-*.csv'
    """

    if filename_pattern[-4:] not in ('.csv', '.tsv'):
        raise ValueError('Extension of `filename_pattern` must match be one of "csv" or "tsv"')

    if filename_pattern.count('*') != 1:
        raise ValueError('`filename_pattern` must have one glob pattern "*"')

    for sheet_name, worksheet in workbook.worksheets.items():
        array = []
        for row in worksheet.rows:
            array.append([cell.value for cell in row.cells])

        pyexcel.save_as(array=array, dest_file_name=filename_pattern.replace('*', '{}').format(sheet_name))


def convert_excel_to_separated_values(filename_excel, filename_pattern_separated_values):
    """ Convert an Excel workbook to a set of csv/tsv files

    Args:
        filename_pattern_separated_values (:obj:`str`): template for file paths, e.g. 'workbook-*.csv'
        filename_excel (:obj:`str`): path to Excel file
    """
    workbook = read_excel(filename_excel)
    write_separated_values(workbook, filename_pattern_separated_values)


def convert_separated_values_to_excel(filename_pattern_separated_values, filename_excel, style=None):
    """ Convert a set of csv/tsv files to an Excel workbook

    Args:
        filename_pattern_separated_values (:obj:`str`): template for file paths, e.g. 'workbook-*.csv'
        filename_excel (:obj:`str`): path to Excel file
        style (:obj:`WorkbookStyle`, optional): workbook style
    """
    workbook = read_separated_values(filename_pattern_separated_values)
    write_excel(workbook, filename_excel, style)


class Workbook(object):
    """ Represents an Excel workbook

    Attributes:
        worksheets (:obj:`OrderedDict`): dictionary of component worksheets
    """

    def __init__(self, worksheets=None):
        """
        Args:
            worksheets (:obj:`OrderedDict`, optional): dictionary of component worksheets
        """
        self.worksheets = worksheets or OrderedDict()


class WorkbookStyle(object):

    def __init__(self, worksheets=None):
        """
        Args:
            worksheets (:obj:`dict`, optional): dictionary of worksheet styles
        """
        self.worksheets = worksheets or {}


class Worksheet(object):
    """ Represents an Excel worksheet

    Attributes:
        rows (:obj: `list` of `Row`): list of rows
    """

    def __init__(self, rows=None):
        """
        Args:
            rows (:obj:`list` of `Row`, optional): list of rows
        """
        self.rows = rows or []


class WorksheetStyle(object):
    """ Represents an Excel worksheet

    Attributes:
        head_rows (:obj: `int`): number of head rows
        head_columns (:obj: `int`): number of head columns
        head_row_font_bold (:obj:`bool`): head row bold
        head_row_fill_pattern (:obj:`str`): head row fill pattern
        head_row_fill_fgcolor (:obj:`str`): head row background color
        row_height (:obj:`float`): row height
    """

    def __init__(self, head_rows=0, head_columns=0, head_row_font_bold=False, head_row_fill_pattern='solid', head_row_fill_fgcolor='', row_height=float('nan')):
        """
        Args:
            head_rows (:obj: `int`, optional): number of head rows
            head_columns (:obj: `int`, optional): number of head columns
            head_row_font_bold (:obj:`bool`, optional): head row bold
            head_row_fill_pattern (:obj:`str`, optional): head row fill pattern
            head_row_fill_fgcolor (:obj:`str`, optional): head row background color
            row_height (:obj:`float`, optional): row height
        """
        self.head_rows = head_rows
        self.head_columns = head_columns
        self.head_row_font_bold = head_row_font_bold
        self.head_row_fill_pattern = head_row_fill_pattern
        self.head_row_fill_fgcolor = head_row_fill_fgcolor
        self.row_height = row_height


class Row(object):
    """ Represents a row of an Excel worksheet

    Attributes:
        cells (:obj: `list` of `Cell`): list of cells
    """

    def __init__(self, cells=None):
        """
        Args:
            cells (:obj:`list` of `Cell`, optional): list of cells
        """
        self.cells = cells or []


class Cell(object):
    """ Represents a cell of an Excel worksheet

    Attributes:
        value (:obj: `object`): value
    """

    def __init__(self, value=None):
        """
        Args:
            value (:obj:`object`, optional): value
        """
        self.value = value
