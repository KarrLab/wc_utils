""" IO utilities

:Author: Jonathan Karr <karr@mssm.edu>
:Author: Arthur Goldberg <Arthur.Goldberg@mssm.edu>
:Date: 2016-11-28
:Copyright: 2016-2018, Karr Lab
:License: MIT
"""

from abc import ABCMeta, abstractmethod
from datetime import datetime
from glob import glob
from itertools import chain
from math import isnan, isinf
from openpyxl import Workbook as XlsWorkbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.colors import Color
from openpyxl.styles.protection import Protection
from openpyxl.utils import get_column_letter
from os.path import basename, dirname, splitext
from shutil import copyfile
from six import integer_types, string_types, with_metaclass
from wc_utils.workbook.core import Workbook, Worksheet, Row
import copy
import enum
import openpyxl.cell.cell
import pyexcel
import xlsxwriter


class Writer(with_metaclass(ABCMeta, object)):
    """ Write data to file(s)

    Attributes:
        path (:obj:`str`): path to file(s)
    """

    def __init__(self, path, title=None, description=None, keywords=None, version=None, language=None, creator=None):
        """
        Args:
            path (:obj:`str`): path to file(s)
            title (:obj:`str`, optional): title
            description (:obj:`str`, optional): description
            keywords (:obj:`str`, optional): keywords
            version (:obj:`str`, optional): version
            language (:obj:`str`, optional): language
            creator (:obj:`str`, optional): creator
        """
        self.path = path
        self.title = title
        self.description = description
        self.keywords = keywords
        self.version = version
        self.language = language
        self.creator = creator

    def run(self, data, style=None, validation=None):
        """ Write workbook to file(s)

        Args:
            data (:obj:`Workbook`): python representation of data; each element must be a string, boolean, integer, float, or NoneType
            style (:obj:`WorkbookStyle`, optional): workbook style
        """
        self.initialize_workbook()

        style = style or WorkbookStyle()
        validation = validation or WorkbookValidation()
        for ws_name, ws_data in data.items():
            ws_style = style.get(ws_name, None)
            ws_validation = validation.get(ws_name, None)
            self.write_worksheet(ws_name, ws_data, style=ws_style, validation=ws_validation)

        self.finalize_workbook()

    @abstractmethod
    def initialize_workbook(self):
        """ Initialize workbook """
        pass  # pragma: no cover

    @abstractmethod
    def write_worksheet(self, sheet_name, data, style=None, validation=None):
        """ Write worksheet to file

        Args:
            sheet_name (:obj:`str`): sheet name
            data (:obj:`Worksheet`): python representation of data; each element must be a string, boolean, integer, float, or NoneType
            style (:obj:`WorksheetStyle`, optional): worksheet style
            validation (:obj:`WorksheetValidation`, optional): worksheet validation
        """
        pass  # pragma: no cover

    @abstractmethod
    def finalize_workbook(self):
        """ Finalize workbook """
        pass  # pragma: no cover


class Reader(with_metaclass(ABCMeta, object)):
    """ Read data from file(s)

    Attributes:
        path (:obj:`str`): path to file(s)
    """

    def __init__(self, path):
        """
        Args:
            path (:obj:`str`): path to file(s)
        """
        self.path = path

    def run(self):
        """ Read data from file(s)

        Returns:
            :obj:`Workbook`: python representation of data
        """
        workbook = self.initialize_workbook()

        names = self.get_sheet_names()
        for name in names:
            workbook[name] = self.read_worksheet(name)

        return workbook

    @abstractmethod
    def initialize_workbook(self):
        """ Initialize workbook

        Returns:
            :obj:`Workbook`: data
        """
        pass  # pragma: no cover

    @abstractmethod
    def get_sheet_names(self):
        """ Get names of sheets contained within path

        Returns:
            obj:`list` of `str`: list of sheet names
        """
        pass  # pragma: no cover

    @abstractmethod
    def read_worksheet(self, sheet_name, ignore_empty_final_rows=True, ignore_empty_final_cols=True):
        """ Read data from file

        Args:
            sheet_name (:obj:`str`): sheet name
            ignore_empty_final_rows (:obj:`bool`, optional): if :obj:`True`, ignore empty final rows
            ignore_empty_final_cols (:obj:`bool`, optional): if :obj:`True`, ignore empty final columns

        Returns:
            :obj:`Worksheet`: data
        """
        pass  # pragma: no cover


class ExcelWriter(Writer):
    """ Write data to Excel file

    Attributes:
        xls_workbook (:obj:`xlsxwriter.Workbook`): Excel workbook
    """

    def __init__(self, path, title=None, description=None, keywords=None, version=None, language=None,
                 creator=None):
        """
        Args:
            path (:obj:`str`): path to file(s)
            title (:obj:`str`, optional): title
            description (:obj:`str`, optional): description
            keywords (:obj:`str`, optional): keywords
            version (:obj:`str`, optional): version
            language (:obj:`str`, optional): language
            creator (:obj:`str`, optional): creator

        Raises:
            :obj:`ValueError`: if file extension is not '.xlsx'
        """
        _, ext = splitext(path)
        if ext != '.xlsx':
            raise ValueError("Extension of path '{}' must be '.xlsx'".format(path))

        super(ExcelWriter, self).__init__(path,
                                          title=title, description=description,
                                          keywords=keywords, version=version, language=language, creator=creator)
        self.xls_workbook = None
        self._worksheet_styles = None

    def initialize_workbook(self):
        """ Initialize workbook """
        # Initialize workbook
        self.xls_workbook = wb = xlsxwriter.Workbook(self.path, {
            'strings_to_numbers': False,
            'strings_to_formulas': False,
            'strings_to_urls': False,
            'nan_inf_to_errors': True,
            'default_date_format': 'yyyy-mm-dd',
        })

        # set metadata
        wb.set_properties({
            'title': self.title,
            'keywords': self.keywords,
        })

        now = datetime.now()
        wb.set_custom_property('description', self.description or '')
        wb.set_custom_property('version', self.version or '')
        wb.set_custom_property('language', self.language or '')
        wb.set_custom_property('creator', self.creator or '')
        wb.set_custom_property('created', now)
        wb.set_custom_property('modified', now)

    def write_worksheet(self, sheet_name, data, style=None, validation=None):
        """ Write worksheet to file

        Args:
            sheet_name (:obj:`str`): sheet name
            data (:obj:`Worksheet`): python representation of data; each element must be a string, boolean, integer, float, or NoneType
            style (:obj:`WorksheetStyle`, optional): worksheet style
            validation (:obj:`WorksheetValidation`, optional): worksheet validation
        """
        xls_worksheet = self.xls_workbook.add_worksheet(sheet_name)

        # data and formatting
        xls_worksheet.protect('', {
            'insert_columns': False,
            'delete_columns': False,

            'insert_rows': True,
            'delete_rows': True,

            'insert_hyperlinks': False,
            'objects': False,
            'scenarios': False,
            'pivot_tables': False,

            'format_cells': False,
            'format_columns': False,
            'format_rows': False,

            'sort': True,
            'autofilter': True,

            'select_locked_cells': True,
            'select_unlocked_cells': True,
        })

        style = style or WorksheetStyle()

        head_format = self.xls_workbook.add_format()
        head_format.set_align('left')
        head_format.set_align('top')
        head_format.set_text_wrap(True)
        head_format.set_font_name(style.font_family)
        head_format.set_font_size(style.font_size)
        head_format.set_bold(True)
        if style.head_row_fill_pattern:
            if style.head_row_fill_pattern == 'solid':
                head_format.set_pattern(1)
            else:
                raise ValueError('Unsupported pattern {}'.format(style.head_row_fill_pattern))
        if style.head_row_fill_fgcolor:
            head_format.set_fg_color('#' + style.head_row_fill_fgcolor)
        head_format.set_locked(True)

        blank_head_format = self.xls_workbook.add_format()
        blank_head_format.set_align('left')
        blank_head_format.set_align('top')
        blank_head_format.set_text_wrap(True)
        blank_head_format.set_font_name(style.font_family)
        blank_head_format.set_font_size(style.font_size)
        blank_head_format.set_bold(True)
        if style.head_row_fill_pattern:
            if style.head_row_fill_pattern == 'solid':
                blank_head_format.set_pattern(1)
            else:  # pragma: no cover # unreachable because error already checked above
                raise ValueError('Unsupported pattern {}'.format(style.head_row_fill_pattern))
        if style.blank_head_fill_fgcolor:
            blank_head_format.set_fg_color('#' + style.blank_head_fill_fgcolor)
        blank_head_format.set_locked(True)

        extra_head_format = self.xls_workbook.add_format()
        extra_head_format.set_align('left')
        extra_head_format.set_align('top')
        extra_head_format.set_text_wrap(True)
        extra_head_format.set_font_name(style.font_family)
        extra_head_format.set_font_size(style.font_size)
        extra_head_format.set_bold(True)
        if style.head_row_fill_pattern:
            if style.head_row_fill_pattern == 'solid':
                extra_head_format.set_pattern(1)
            else:  # pragma: no cover # unreachable because error already checked above
                raise ValueError('Unsupported pattern {}'.format(style.head_row_fill_pattern))
        if style.head_row_fill_fgcolor:
            extra_head_format.set_fg_color('#' + style.head_row_fill_fgcolor)
        extra_head_format.set_locked(False)

        merged_head_format = self.xls_workbook.add_format()
        merged_head_format.set_align('center')
        merged_head_format.set_align('top')
        merged_head_format.set_text_wrap(True)
        merged_head_format.set_font_name(style.font_family)
        merged_head_format.set_font_size(style.font_size)
        merged_head_format.set_bold(True)
        if style.head_row_fill_pattern:
            if style.head_row_fill_pattern == 'solid':
                merged_head_format.set_pattern(1)
            else:  # pragma: no cover # unreachable because error already checked above
                raise ValueError('Unsupported pattern {}'.format(style.head_row_fill_pattern))
        if style.merged_head_fill_fgcolor:
            merged_head_format.set_fg_color('#' + style.merged_head_fill_fgcolor)
        merged_head_format.set_locked(True)

        body_format = self.xls_workbook.add_format()
        body_format.set_align('left')
        body_format.set_align('top')
        body_format.set_text_wrap(True)
        body_format.set_font_name(style.font_family)
        body_format.set_font_size(style.font_size)
        body_format.set_bold(False)
        body_format.set_locked(False)

        merge_body_format = self.xls_workbook.add_format()
        merge_body_format.set_align('center')
        merge_body_format.set_align('vcenter')
        merge_body_format.set_text_wrap(True)
        merge_body_format.set_font_name(style.font_family)
        merge_body_format.set_font_size(style.font_size)
        merge_body_format.set_bold(False)
        merge_body_format.set_locked(False)

        n_rows = len(data)
        if data:
            n_cols = max(len(row) for row in data)
        else:
            n_cols = 0
        frozen_rows = style.head_rows
        frozen_columns = style.head_columns
        row_height = style.row_height
        col_width = style.col_width

        # format rows
        if isnan(row_height):
            default_row_height = None
        else:
            default_row_height = row_height
        hide_unused_rows = not isinf(style.extra_rows)
        xls_worksheet.set_default_row(default_row_height, hide_unused_rows)

        # format columns
        if not isnan(col_width) and n_cols >= 1 and not isinf(style.extra_columns):
            result = xls_worksheet.set_column(0, n_cols - 1, width=col_width, options={'hidden': False})
            assert result == 0, result

        # hyperlinks
        for hyperlink in style.hyperlinks:
            result = xls_worksheet.write_url(hyperlink.i_row, hyperlink.i_col, hyperlink.url, tip=hyperlink.tip)
            assert result == 0

        # write data
        for i_row, row in enumerate(data):
            for i_col, value in enumerate(row):
                if i_row < frozen_rows or i_col < frozen_columns:
                    if value is None or value == '':
                        format = blank_head_format
                    else:
                        format = head_format
                else:
                    format = body_format

                self.write_cell(xls_worksheet, sheet_name, i_row, i_col, value, format)

            if not isnan(row_height) and not isinf(style.extra_rows):
                result = xls_worksheet.set_row(i_row, options={'hidden': False})
                assert result in [0, None], result

        # format extra columns
        if isinf(style.extra_columns):
            extra_columns = min(100, 2**14 - n_cols)
        else:
            extra_columns = style.extra_columns
            result = xls_worksheet.set_column(n_cols + style.extra_columns, 2**14 - 1,
                                              options={'hidden': True})
            assert result == 0, result

        for i_row in range(n_rows):
            for i_col in range(n_cols, n_cols + extra_columns):
                if i_row < frozen_rows or i_col < frozen_columns:
                    format = extra_head_format
                else:
                    format = body_format
                result = xls_worksheet.write_blank(i_row, i_col, None, format)
                assert result == 0, result

        # format extra rows
        if isinf(style.extra_rows):
            extra_rows = min(100, 2**20 - n_rows)
        else:
            extra_rows = style.extra_rows
            for i_row in range(n_rows, n_rows + style.extra_rows):
                result = xls_worksheet.set_row(i_row, options={'hidden': False})
                assert result in [0, None], result

                for i_col in range(n_cols + extra_columns):
                    if i_row < frozen_rows or i_col < frozen_columns:
                        format = extra_head_format
                    else:
                        format = body_format
                    result = xls_worksheet.write_blank(i_row, i_col, None, format)
                    assert result == 0, result

        # merge ranges
        for row_start, col_start, row_end, col_end in style.merge_ranges:
            # get data
            value = set()
            for i_row in range(row_start, row_end + 1):
                for i_col in range(col_start, col_end + 1):
                    if data[i_row][i_col] is not None:
                        value.add(data[i_row][i_col])
            if len(value) == 0:
                value = None
            elif len(value) == 1:
                value = list(value)[0]
            else:
                raise ValueError('Merge range {}{}:{}{} with values {{"{}"}} can have at most 1 value'.format(
                    get_column_letter(col_start + 1), row_start + 1,
                    get_column_letter(col_end + 1), row_end + 1,
                    '", "'.join(str(v) for v in value)))

            if row_start <= frozen_rows or col_start <= frozen_columns:
                format = merged_head_format
            else:
                format = merge_body_format
            xls_worksheet.merge_range(row_start, col_start, row_end, col_end, None)
            self.write_cell(xls_worksheet, sheet_name, row_start, col_start, value, format)

        # validation
        if validation:
            validation.apply(xls_worksheet,
                             frozen_rows, frozen_columns,
                             n_rows + extra_rows - 1, n_cols + extra_columns - 1)

        # freeze panes
        xls_worksheet.freeze_panes(frozen_rows, frozen_columns)

        # auto filter
        if style.auto_filter and n_cols > 0 and n_cols > 0 and frozen_rows > 0:
            xls_worksheet.autofilter(frozen_rows - 1, 0, n_rows - 1, n_cols - 1)

    def write_cell(self, xls_worksheet, sheet_name, i_row, i_col, value, format):
        """ Write a value to a cell

        Args:
            xls_worksheet (:obj:`xlsxwriter.Worksheet`): Excel worksheet
            sheet_name (:obj:`str`): sheet name
            i_row (:obj:`int`): row of cell to write
            i_col (:obj:`int`): column of cell to write
            value (:obj:`object`): value to write
            format (:obj:`xlsxwriter.Format`): format for the cell
        """
        if value is None or value == '':
            result = xls_worksheet.write_blank(i_row, i_col, value, format)
        elif isinstance(value, string_types):
            result = xls_worksheet.write_string(i_row, i_col, value, format)
        elif isinstance(value, bool):
            result = xls_worksheet.write_boolean(i_row, i_col, value, format)
        elif isinstance(value, integer_types):
            result = xls_worksheet.write_number(i_row, i_col, float(value), format)
        elif isinstance(value, float):
            result = xls_worksheet.write_number(i_row, i_col, value, format)
        else:
            raise ValueError('Unsupported type {} at {}:{}:{}{}'.format(
                value.__class__.__name__,
                self.path, sheet_name, get_column_letter(i_col + 1), i_row + 1))
        assert result == 0, 'Error code {} when writing "{}" to worksheet "{}"'.format(
            result, value, sheet_name)

    def finalize_workbook(self):
        """ Finalize workbook """
        self.xls_workbook.close()


class ExcelReader(Reader):
    """ Read data from Excel file

    Attributes:
        xls_workbook (:obj:`Workbook`): Excel workbook
    """

    def __init__(self, path):
        """
        Args:
            path (:obj:`str`): path to file(s)

        Raises:
            :obj:`ValueError`: if file extension is not '.xlsx'
        """
        _, ext = splitext(path)
        if ext != '.xlsx':
            raise ValueError("Extension of path '{}' must be '.xlsx'".format(path))
        super(ExcelReader, self).__init__(path)
        self.xls_workbook = None

    def initialize_workbook(self):
        """ Initialize workbook

        Returns:
            :obj:`Workbook`: data
        """
        self.xls_workbook = load_workbook(filename=self.path)
        return Workbook()

    def get_sheet_names(self):
        """ Get names of sheets contained within path

        Returns:
            obj:`list` of `str`: list of sheet names
        """
        return self.xls_workbook.sheetnames

    def read_worksheet(self, sheet_name, ignore_empty_final_rows=True, ignore_empty_final_cols=True):
        """ Read data from Excel worksheet

        Args:
            sheet_name (:obj:`str`): sheet name
            ignore_empty_final_rows (:obj:`bool`, optional): if :obj:`True`, ignore empty final rows
            ignore_empty_final_cols (:obj:`bool`, optional): if :obj:`True`, ignore empty final columns

        Returns:
            :obj:`Worksheet`: data

        Raises:
            :obj:`ValueError`:
        """
        xls_worksheet = self.xls_workbook[sheet_name]
        worksheet = Worksheet()

        max_row = xls_worksheet.max_row
        max_col = xls_worksheet.max_column

        if ignore_empty_final_rows:
            real_max_row = None
            for i_row in range(xls_worksheet.max_row, 0, -1):
                for i_col in range(1, max_col + 1):
                    value = self.read_cell(sheet_name, xls_worksheet, i_row, i_col)
                    if value not in (None, ''):
                        real_max_row = i_row
                        break
                if real_max_row is not None:
                    break
            if real_max_row is not None:
                max_row = real_max_row
            else:
                max_row = 0

        if ignore_empty_final_cols:
            real_max_col = None
            for i_col in range(max_col, 0, -1):
                for i_row in range(1, max_row + 1):
                    value = self.read_cell(sheet_name, xls_worksheet, i_row, i_col)
                    if value not in (None, ''):
                        real_max_col = i_col
                        break
                if real_max_col is not None:
                    break
            if real_max_col is not None:
                max_col = real_max_col
            else:
                max_col = 0

        for i_row in range(1, max_row + 1):
            row = Row()
            worksheet.append(row)
            for i_col in range(1, max_col + 1):
                value = self.read_cell(sheet_name, xls_worksheet, i_row, i_col)
                row.append(value)

        for cell in xls_worksheet.merged_cells.ranges:
            value = worksheet[cell.min_row-1][cell.min_col-1]
            for i_row in range(cell.min_row-1, cell.max_row):
                for i_col in range(cell.min_col-1, cell.max_col):
                    worksheet[i_row][i_col] = value

        return worksheet

    def read_cell(self, sheet_name,  xls_worksheet, i_row, i_col):
        """ Read the value of a cell

        Args:
            sheet_name (:obj:`str`): worksheet name
            xls_worksheet (:obj:`openpyxl.Worksheet`): worksheet
            i_row (:obj:`int`): row number
            i_col (:obj:`int`): column number

        Returns:
            :obj:`object`: value of cell
        """
        cell = xls_worksheet.cell(row=i_row, column=i_col)

        if cell.data_type in (openpyxl.cell.cell.TYPE_STRING, openpyxl.cell.cell.TYPE_INLINE,
                              openpyxl.cell.cell.TYPE_NUMERIC, openpyxl.cell.cell.TYPE_NULL,
                              openpyxl.cell.cell.TYPE_BOOL):
            value = cell.value
        elif cell.data_type == openpyxl.cell.cell.TYPE_ERROR:
            raise ValueError('Errors are not supported: {}:{}:{}{}'.format(self.path, sheet_name,
                                                                           get_column_letter(i_col), i_row))
        elif cell.data_type in (openpyxl.cell.cell.TYPE_FORMULA,
                                openpyxl.cell.cell.TYPE_FORMULA_CACHE_STRING):
            if cell.value in ['=FALSE()', '=FALSE']:
                value = False
            elif cell.value in ['=TRUE()', '=TRUE']:
                value = True
            else:
                raise ValueError('Formula are not supported: {}:{}:{}{}'.format(
                    self.path, sheet_name, get_column_letter(i_col), i_row))
        else:
            raise ValueError('Unsupported data type: {} at {}:{}:{}{}'.format(
                cell.data_type, self.path, sheet_name, get_column_letter(i_col), i_row))  # pragma: no cover # unreachable

        return value


class SeparatedValuesWriter(Writer):
    """ Write data to csv/tsv file(s) """

    def __init__(self, path, title=None, description=None, keywords=None, version=None, language=None, creator=None):
        """
        Args:
            path (:obj:`str`): path to file(s)
            title (:obj:`str`, optional): title
            description (:obj:`str`, optional): description
            keywords (:obj:`str`, optional): keywords
            version (:obj:`str`, optional): version
            language (:obj:`str`, optional): language
            creator (:obj:`str`, optional): creator

        Raises:
            :obj:`ValueError`: if file extension is not '.csv' or '.tsv' or if file name pattern
                doesn't contain exactly one glob
        """
        _, ext = splitext(path)
        if ext not in ('.csv', '.tsv'):
            raise ValueError("Extension of path '{}' must be one of '.csv' or '.tsv'".format(
                path))

        if '*' in dirname(path):
            raise ValueError("path '{}' cannot have glob pattern '*' in its directory name".format(
                path))

        if basename(path).count('*') != 1:
            raise ValueError("path '{}' must have one glob pattern '*' in its base name".format(
                path))

        super(SeparatedValuesWriter, self).__init__(path,
                                                    title=title, description=description,
                                                    keywords=keywords, version=version, language=language, creator=creator)

    def initialize_workbook(self):
        """ Initialize workbook """
        pass

    def write_worksheet(self, sheet_name, data, style=None, validation=None):
        """ Write worksheet to file

        Args:
            sheet_name (:obj:`str`): sheet name
            data (:obj:`Worksheet`): python representation of data; each element must be a string, boolean, integer, float, or NoneType
            style (:obj:`WorksheetStyle`, optional): worksheet style
            validation (:obj:`WorksheetValidation`, optional): worksheet validation
        """
        pyexcel.save_as(array=data, dest_file_name=self.path.replace('*', '{}').format(sheet_name))

    def finalize_workbook(self):
        """ Finalize workbook """
        pass


class SeparatedValuesReader(Reader):
    """ Read data from csv/tsv file(s) """

    def __init__(self, path):
        """
        Args:
            path (:obj:`str`): path to file(s)

        Raises:
            :obj:`ValueError`: if file extension is not '.csv' or '.tsv' or if file name pattern
                doesn't contain exactly one glob
        """
        _, ext = splitext(path)
        if ext not in ('.csv', '.tsv'):
            raise ValueError("Extension of path '{}' must be one of '.csv' or '.tsv'".format(
                path))

        if '*' in dirname(path):
            raise ValueError("path '{}' cannot have glob pattern '*' in its directory name".format(
                path))

        if basename(path).count('*') != 1:
            raise ValueError("path '{}' must have one glob pattern '*' in its base name".format(
                path))

        super(SeparatedValuesReader, self).__init__(path)

    def initialize_workbook(self):
        """ Initialize workbook

        Returns:
            :obj:`Workbook`: data
        """
        return Workbook()

    def get_sheet_names(self):
        """ Get names of files contained within path glob

        Returns:
            obj:`list` of `str`: list of file names

        Raises:
            :obj:`ValueError`: if glob does not find any matching files
        """
        i_glob = self.path.find('*')
        names = []
        for filename in glob(self.path):
            names.append(filename[i_glob:i_glob + len(filename) - len(self.path) + 1])
        if not names:
            raise ValueError("glob of path '{}' does not match any files".format(self.path))
        return names

    def read_worksheet(self, sheet_name, ignore_empty_final_rows=True, ignore_empty_final_cols=True):
        """ Read data from file

        Args:
            sheet_name (:obj:`str`): sheet name
            ignore_empty_final_rows (:obj:`bool`, optional): if :obj:`True`, ignore empty final rows
            ignore_empty_final_cols (:obj:`bool`, optional): if :obj:`True`, ignore empty final columns

        Returns:
            :obj:`Worksheet`: data
        """
        worksheet = Worksheet()
        # todo: skip_empty_rows=False is the default for pyexcel-io v >= 0.3.2
        # when it's available on pypi, set pyexcel>=0.4.0  pyexcel-io>=0.3.2 & remove skip_empty_rows option
        sv_worksheet = pyexcel.get_sheet(file_name=self.path.replace('*', '{}').format(sheet_name),
                                         skip_empty_rows=False)

        rows = list(sv_worksheet.rows())
        max_row = len(rows)
        if max_row:
            max_col = len(rows[0])
        else:
            max_col = 0

        if ignore_empty_final_rows:
            real_max_row = None
            for i_row, row in enumerate(reversed(rows)):
                for i_col, cell in enumerate(row):
                    value = self.read_cell(cell)
                    if value not in (None, ''):
                        real_max_row = max_row - i_row
                        break
                if real_max_row is not None:
                    break
            if real_max_row is not None:
                max_row = real_max_row

        if ignore_empty_final_cols:
            real_max_col = None
            for i_col in range(max_col - 1, 0, -1):
                for i_row in range(0, max_row):
                    cell = rows[i_row][i_col]
                    value = self.read_cell(cell)
                    if value not in (None, ''):
                        real_max_col = i_col + 1
                        break
                if real_max_col is not None:
                    break
            if real_max_col is not None:
                max_col = real_max_col

        for sv_row in rows[0:max_row]:
            row = Row()
            worksheet.append(row)
            for sv_cell in sv_row[0:max_col]:
                row.append(self.read_cell(sv_cell))

        return worksheet

    def read_cell(self, value):
        """ Read the value of a cell

        Args:
            value (:obj:`object`): value

        Returns:
            :obj:`object`: value
        """
        if value == '':
            value = None
        elif value == 'True':
            value = True
        elif value == 'False':
            value = False
        return value


def get_writer(extension):
    """ Get writer

    Args:
        extension (:obj:`str`): extension

    Returns:
        :obj:`class`: writer class

    Raises:
        :obj:`ValueError`: if extension is not one of ".xlsx", ".csv", or ".tsv"
    """
    if extension == '.xlsx':
        return ExcelWriter
    elif extension in ['.csv', '.tsv']:
        return SeparatedValuesWriter
    else:
        raise ValueError('Extension must be one of ".xlsx", ".csv", or ".tsv"')


def get_reader(extension):
    """ Get reader

    Args:
        extension (:obj:`str`): extension

    Returns:
        :obj:`class`: reader class

    Raises:
        :obj:`ValueError`: if extension is not one of ".xlsx", ".csv", or ".tsv"
    """
    if extension == '.xlsx':
        return ExcelReader
    elif extension in ['.csv', '.tsv']:
        return SeparatedValuesReader
    else:
        raise ValueError('Extension must be one of ".xlsx", ".csv", or ".tsv"')


def write(path, workbook,
          title=None, description=None, keywords=None, version=None, language=None, creator=None,
          style=None):
    """ Write data to Excel (.xlsx) file or collection of comma separated (.csv) or tab separated (.tsv) file(s)

    Args:
        path (:obj:`str`): path to file(s)
        workbook (:obj:`Workbook`): python representation of data; each element must be a string, boolean, integer, float, or NoneType
        title (:obj:`str`, optional): title
        description (:obj:`str`, optional): description
        keywords (:obj:`str`, optional): keywords
        version (:obj:`str`, optional): version
        language (:obj:`str`, optional): language
        creator (:obj:`str`, optional): creator
        style (:obj:`WorkbookStyle`, optional): workbook style
    """
    # check extensions are valid
    _, ext = splitext(path)
    writer_cls = get_writer(ext)

    writer = writer_cls(path,
                        title=title, description=description, keywords=keywords,
                        version=version, language=language, creator=creator)
    writer.run(workbook, style=style)


def read(path):
    """ Read data from Excel (.xlsx) file or collection of comma separated (.csv) or tab separated (.tsv) file(s)

    Args:
        path (:obj:`str`): path to file(s)

    Returns:
        :obj:`Workbook`: python representation of data
    """
    # check extensions are valid
    _, ext = splitext(path)
    reader_cls = get_reader(ext)
    reader = reader_cls(path)
    return reader.run()


def convert(source, destination, worksheet_order=None, style=None, ignore_extra_sheets=True):
    """ Convert among Excel (.xlsx), comma separated (.csv), and tab separated formats (.tsv)

    Args:
        source (:obj:`str`): path to source file
        destination (:obj:`str`): path to save converted file
        worksheet_order (:obj:`list` of :obj:`str`): worksheet order
        style (:obj:`WorkbookStyle`, optional): workbook style for Excel
        ignore_extra_sheets (:obj:`bool`, optional): true/false should extra sheets in worksheet_order be ignored or should an error be thrown

    Raises:
        :obj:`ValueError`: if file extensions are not supported or file names are equal
    """
    # check source != destination
    if source == destination:
        raise ValueError('Source and destination names must be different')

    # check extensions are valid
    _, ext_src = splitext(source)
    _, ext_dst = splitext(destination)

    if ext_src not in ['.xlsx', '.csv', '.tsv']:
        raise ValueError('Source extension must be one of ".xlsx", ".csv", or ".tsv"')

    if ext_dst not in ['.xlsx', '.csv', '.tsv']:
        raise ValueError('Destination extension must be one of ".xlsx", ".csv", or ".tsv"')

    # if extensions are the same, copy file(s)
    if ext_src == ext_dst and (worksheet_order is None or ext_src != '.xlsx'):
        if ext_src == '.xlsx':
            copyfile(source, destination)
        else:
            i_glob = source.find('*')
            dst_format = destination.replace('*', '{}')
            if not list(glob(source)):
                raise ValueError("glob of path '{}' does not match any files".format(source))
            for filename in glob(source):
                sheet_name = filename[i_glob:i_glob + len(filename) - len(source) + 1]
                copyfile(filename, dst_format.format(sheet_name))
        return

    # read, convert, and write
    workbook = read(source)

    ordered_workbook = Workbook()
    worksheet_order = worksheet_order or []
    if not ignore_extra_sheets:
        difference = set(worksheet_order) - set(workbook.keys())
        if difference:
            raise ValueError("source '{}' missing worksheets: '{}'".format(source, difference))

    for worksheet in chain(worksheet_order, workbook.keys()):
        if worksheet in workbook:
            ordered_workbook[worksheet] = workbook[worksheet]

    write(destination, ordered_workbook, style=style)


class WorkbookStyle(dict):
    """ Workbook style: dictionary of worksheet styles """
    pass


class WorksheetStyle(object):
    """ Worksheet style

    Attributes:
        head_rows (:obj:`int`): number of head rows
        head_columns (:obj:`int`): number of head columns
        head_row_font_bold (:obj:`bool`): head row bold
        head_row_fill_pattern (:obj:`str`): head row fill pattern
        head_row_fill_fgcolor (:obj:`str`): head row background color
        blank_head_fill_fgcolor (:obj:`str`): background color of blank header cells
        merged_head_fill_fgcolor (:obj:`str`): background color of merged header cells
        extra_rows (:obj:`float`): number of additional rows to show
        extra_columns (:obj:`float`): number of additional columns to show
        font_family (:obj:`str`): font family
        font_size (:obj:`float`): font size
        row_height (:obj:`float`): row height
        col_width (:obj:`float`): column width
        auto_filter (:obj:`bool`): whether or not to activate auto filters for row
        merge_ranges (:obj:`list` of :obj:`tuple` of :obj:`int`): list of tuples of the start row, start column, end row, and end column (0-based) 
            of each range to merge
        hyperlinks (:obj:`list` of :obj:`Hyperlink`): list of hyperlinks
    """

    def __init__(self, head_rows=0, head_columns=0, head_row_font_bold=False,
                 head_row_fill_pattern='solid', head_row_fill_fgcolor='', blank_head_fill_fgcolor='', merged_head_fill_fgcolor='',
                 extra_rows=float('inf'), extra_columns=float('inf'),
                 font_family='Arial', font_size=11.,
                 row_height=15., col_width=15.,
                 auto_filter=True, merge_ranges=None, hyperlinks=None):
        """
        Args:
            head_rows (:obj:`int`, optional): number of head rows
            head_columns (:obj:`int`, optional): number of head columns
            head_row_font_bold (:obj:`bool`, optional): head row bold
            head_row_fill_pattern (:obj:`str`, optional): head row fill pattern
            head_row_fill_fgcolor (:obj:`str`, optional): head row background color
            blank_head_fill_fgcolor (:obj:`str`, optional): background color of blank header cells
            merged_head_fill_fgcolor (:obj:`str`, optional): background color of merged header cells
            extra_rows (:obj:`float`, optional): number of additional rows to show
            extra_columns (:obj:`float`, optional): number of additional columns to show
            font_family (:obj:`str`, optional): font family
            font_size (:obj:`float`, optional): font size
            row_height (:obj:`float`, optional): row height
            col_width (:obj:`float`, optional): column width
            auto_filter (:obj:`bool`, optional): whether or not to activate auto filters for row
            merge_ranges (:obj:`list` of :obj:`tuple` of :obj:`int`, optional): list of tuples of the start row, start column, end row,
                and end column (0-based) of each range to merge
            hyperlinks (:obj:`list` of :obj:`Hyperlink`, optional): list of hyperlinks
        """
        self.head_rows = head_rows
        self.head_columns = head_columns
        self.head_row_font_bold = head_row_font_bold
        self.head_row_fill_pattern = head_row_fill_pattern
        self.head_row_fill_fgcolor = head_row_fill_fgcolor
        self.blank_head_fill_fgcolor = blank_head_fill_fgcolor
        self.merged_head_fill_fgcolor = merged_head_fill_fgcolor
        self.extra_rows = extra_rows
        self.extra_columns = extra_columns
        self.font_family = font_family
        self.font_size = font_size
        self.row_height = row_height
        self.col_width = col_width
        self.auto_filter = auto_filter
        self.merge_ranges = merge_ranges or []
        self.hyperlinks = hyperlinks or []


class Hyperlink(object):
    """ Hyperlink from a cell

    Attributes:
        i_row (:obj:`int`): row
        i_row (:obj:`col`): column
        url (:obj:`str`): URL
        tip (:obj:`str`): text of tooltip
    """

    def __init__(self, i_row, i_col, url, tip=None):
        """
        Args:
            i_row (:obj:`int`): row
            i_row (:obj:`col`): column
            url (:obj:`str`): URL
            tip (:obj:`str`, optional): text of tooltip
        """
        self.i_row = i_row
        self.i_col = i_col
        self.url = url
        self.tip = tip


class WorkbookValidation(dict):
    """ Workbook validation: dictionary of worksheet validations """
    pass


class WorksheetValidationOrientation(int, enum.Enum):
    """ Worksheet validation orientation """
    row = 1
    column = 2


class WorksheetValidation(object):
    """ List of field validations

    Attributes:
        orientation (:obj:`str`): row or col
        fields (:obj:`list` of :obj:`FieldValidation`): field validations
    """

    def __init__(self, orientation=WorksheetValidationOrientation.row, fields=None):
        """
        Args:
            orientation (:obj:`str`, optional): row or col
            fields (:obj:`list` of :obj:`FieldValidation`, optional): field validations
        """
        self.orientation = orientation
        self.fields = fields

    def apply(self, ws, first_row, first_col, last_row, last_col):
        """ Apply validation to worksheet

        Args:
            ws (:obj:`xlsxwriter.Worksheet`): worksheet
            first_row (:obj:`int`): first row
            first_col (:obj:`int`): first column
            last_row (:obj:`int`): last row
            last_col (:obj:`int`): last column
        """
        for i_field, field in enumerate(self.fields):
            if field:
                if self.orientation == WorksheetValidationOrientation.row:
                    field.apply_help_comment(ws, first_row - 1, i_field)
                    field.apply_validation(ws, first_row, i_field, last_row, i_field)
                else:
                    field.apply_help_comment(ws, i_field, first_col - 1)
                    field.apply_validation(ws, i_field, first_col, i_field, last_col)


class FieldValidationType(int, enum.Enum):
    """ Field validation type """
    integer = 1
    decimal = 2
    list = 3
    date = 4
    time = 5
    length = 6
    custom = 7
    any = 8


FieldValidationCriterion = enum.Enum('FieldValidationCriterion', type=str, names=[
    ('between', 'between'),
    ('not between', 'not between'),
    ('equal to', '=='),
    ('not equal to', '!='),
    ('greater than', '>'),
    ('less than', '<'),
    ('greater than or equal to', '>='),
    ('less than or equal to', '<='),
    ('==', '=='),
    ('!=', '!='),
    ('>', '>'),
    ('<', '<'),
    ('>=', '>='),
    ('<=', '<='),
])


class FieldValidationErrorType(int, enum.Enum):
    """ Type of error dialog to display """
    stop = 1
    warning = 2
    information = 3


class FieldValidation(object):
    """ Validation for row- or column-oriented field

    Attributes:
        input_title (:obj:`str`): title of input dialog box
        input_message (:obj:`str`): message in input dialog box
        show_input (:obj:`bool`): if :obj:`True`, show input dialog box
        type (:obj:`FieldValidationType`): type of validation
        criterion (:obj:`FieldValidationCriterion`): validation criterion
        allowed_scalar_value (:obj:`bool`, :obj:`int`, :obj:`float`, or :obj:`str`): allowable scalar value
        minimum_scalar_value (:obj:`int` or :obj:`float`): minimum allowable value
        maximum_scalar_value (:obj:`int` or :obj:`float`): maximum allowable value
        allowed_list_values (:obj:`str` or :obj:`list` of :obj:`str`): allowable list values
        show_dropdown (:obj:`bool`): if :obj:`True`, show dropdown menu for list validations
        ignore_blank (:obj:`bool`): if :obj:`True`, don't validate blank cells
        error_type (:obj:`FieldErrorType`): type of error dialog to display
        error_title (:obj:`str`): title of error dialog box
        error_message (:obj:`str`): message in error dialog box
        show_error (:obj:`bool`): if :obj:`True`, show error dialog box
    """

    def __init__(self, input_title='', input_message='', show_input=True,
                 type=FieldValidationType.any, criterion=None, allowed_scalar_value=None,
                 minimum_scalar_value=None, maximum_scalar_value=None, allowed_list_values=None,
                 show_dropdown=True, ignore_blank=True,
                 error_type=FieldValidationErrorType.warning, error_title='', error_message='',  show_error=True):
        """
        Args:
            input_title (:obj:`str`, optional): title of input dialog box
            input_message (:obj:`str`, optional): message in input dialog box
            show_input (:obj:`bool`, optional): if :obj:`True`, show input dialog box
            type (:obj:`FieldValidationType`, optional: type of validation
            criterion (:obj:`FieldValidationCriterion`, optional): validation criterion
            allowed_scalar_value (:obj:`bool`, :obj:`int`, :obj:`float`, or :obj:`str`, optional): allowable scalar value
            minimum_scalar_value (:obj:`int` or :obj:`float`, optional): minimum allowable value
            maximum_scalar_value (:obj:`int` or :obj:`float`, optional): maximum allowable value
            allowed_list_values (:obj:`str` or :obj:`list`, optional): allowable list values
            show_dropdown (:obj:`bool`, optional): if :obj:`True`, show dropdown menu for list validations
            ignore_blank (:obj:`bool`, optional): if :obj:`True`, don't validate blank cells
            error_type (:obj:`FieldErrorType`, optional): type of error dialog to display
            error_title (:obj:`str`, optional): title of error dialog box
            error_message (:obj:`str`, optional): message in error dialog box
            show_error (:obj:`bool`, optional): if :obj:`True`, show error dialog box
        """
        self.input_title = input_title
        self.input_message = input_message
        self.show_input = show_input
        self.type = type
        self.criterion = criterion
        self.allowed_scalar_value = allowed_scalar_value
        self.minimum_scalar_value = minimum_scalar_value
        self.maximum_scalar_value = maximum_scalar_value
        self.allowed_list_values = allowed_list_values
        self.show_dropdown = show_dropdown
        self.ignore_blank = ignore_blank
        self.error_type = error_type
        self.error_title = error_title
        self.error_message = error_message
        self.show_error = show_error

    def apply_help_comment(self, ws, i_row, i_col):
        """ Apply help comment to cell 

        Args:
            ws (:obj:`xlsxwriter.Worksheet`): worksheet
            i_row (:obj:`int`): row        
            i_col (:obj:`int`): column
        """
        ws.write_comment(i_row, i_col, self.input_message, {
            'author': None,
            'visible': False,
            'font_name': 'Arial',
            'font_size': 10,
            'width': 300,  # pixels
        })

    def apply_validation(self, ws, first_row, first_col, last_row, last_col):
        """ Apply validation to cells

        Args:
            ws (:obj:`xlsxwriter.Worksheet`): worksheet
            first_row (:obj:`int`): first row
            first_col (:obj:`int`): first column
            last_row (:obj:`int`): last row
            last_col (:obj:`int`): last column
        """
        ws.data_validation(first_row, first_col, last_row, last_col, self.get_options())

    def get_options(self):
        """ Get options for :obj:`xlsxwriter.Worksheet.data_validation`

        Returns
            :obj:`dict`: dictonary of options for :obj:`xlsxwriter.Worksheet.data_validation`
        """
        options = {}

        # input dialog
        if self.input_title:
            if len(self.input_title) > 32:
                options['input_title'] = self.input_title[0:32-4] + ' ...'
            else:
                options['input_title'] = self.input_title
        if self.input_message:
            if len(self.input_message) > 255:
                options['input_message'] = self.input_message[0:255-4] + ' ...'
            else:
                options['input_message'] = self.input_message
        options['show_input'] = self.show_input

        # validation
        options['validate'] = self.type.name

        if self.criterion:
            options['criteria'] = self.criterion.value

        if self.allowed_scalar_value:
            options['value'] = self.allowed_scalar_value

        if self.minimum_scalar_value:
            options['minimum'] = self.minimum_scalar_value

        if self.maximum_scalar_value:
            options['maximum'] = self.maximum_scalar_value

        if self.allowed_list_values:
            options['source'] = self.allowed_list_values

        options['dropdown'] = self.show_dropdown
        options['ignore_blank'] = self.ignore_blank

        # error dialog
        options['error_type'] = self.error_type.name
        if self.error_title:
            if len(self.error_title) > 32:
                options['error_title'] = self.error_title[0:32-4] + ' ...'
            else:
                options['error_title'] = self.error_title
        if self.error_message:
            if len(self.error_message) > 255:
                options['error_message'] = self.error_message[0:255-4] + ' ...'
            else:
                options['error_message'] = self.error_message
        options['show_error'] = self.show_error

        return options
